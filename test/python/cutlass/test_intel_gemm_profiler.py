#################################################################################################
# Copyright (C) 2026 Intel Corporation, All rights reserved.
# SPDX-License-Identifier: BSD-3-Clause
#################################################################################################

import importlib.util
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


def load_module():
    repo_root = Path(__file__).resolve().parents[3]
    module_path = repo_root / "test" / "benchmarks" / "intel_gemm_profiler.py"
    spec = importlib.util.spec_from_file_location("intel_gemm_profiler", module_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


profiler = load_module()


class TestIntelGemmProfiler(unittest.TestCase):
    def test_generate_candidate_space_uses_catalog_level0(self):
        shapes = profiler.default_shapes("bf16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()

        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles)

        self.assertEqual(candidate_space["device_arch"], "bmg")
        self.assertEqual(candidate_space["kernel_catalog"]["catalog_version"], "level0-seed-catalog")
        self.assertEqual(
            candidate_space["search_runtime_schema"]["compile_time_dimensions"][0],
            "dtype_a",
        )
        self.assertEqual(len(candidate_space["candidates"]), 8)
        candidate_ids = {candidate["candidate_id"] for candidate in candidate_space["candidates"]}
        self.assertIn("rcr_bf16bf16f32_tm8_tn128_tk32_sg1x4_st2_sk1", candidate_ids)
        self.assertIn("rcr_bf16bf16f32_tm256_tn256_tk32_sg8x4_st2_sk1", candidate_ids)
        self.assertIn("rcr_bf16bf16f32_tm128_tn256_tk32_sg4x4_st2_sk1", candidate_ids)
        self.assertFalse(any(candidate["split_k"] > 1 for candidate in candidate_space["candidates"]))
        self.assertTrue(all(candidate["filters_applied"][0] == "kernel_catalog" for candidate in candidate_space["candidates"]))
        self.assertTrue(all(candidate["grf_mode"] == 256 for candidate in candidate_space["candidates"]))

        probe_candidate_space = profiler.generate_candidate_space(
            shapes, constraints, profiles, allowed_runners=("benchmark", "streamk_example")
        )
        self.assertEqual(len(probe_candidate_space["candidates"]), 11)
        streamk = next(
            candidate
            for candidate in probe_candidate_space["candidates"]
            if candidate["streamk_mode"] == "streamk"
        )
        data_parallel = next(
            candidate
            for candidate in probe_candidate_space["candidates"]
            if candidate["streamk_mode"] == "data_parallel"
        )
        splitk = next(candidate for candidate in probe_candidate_space["candidates"] if candidate["split_k"] == 2)
        self.assertEqual(streamk["runner"], "streamk_example")
        self.assertEqual(data_parallel["runner"], "streamk_example")
        self.assertEqual(splitk["runner"], "streamk_example")
        self.assertEqual(splitk["benchmark_target"], "03_bmg_gemm_streamk")

    def test_build_kernel_catalog_includes_runtime_metadata(self):
        catalog = profiler.build_kernel_catalog(dtypes=["bf16"], allowed_runners=("benchmark", "streamk_example"))

        self.assertEqual(catalog["catalog_version"], "level0-seed-catalog")
        self.assertEqual(catalog["search_runtime_schema"]["runtime_dimensions"][0], "shape_id")
        self.assertEqual(len(catalog["kernels"]), 12)
        splitk = next(entry for entry in catalog["kernels"] if entry["split_k"] == 2)
        data_parallel = next(entry for entry in catalog["kernels"] if entry["streamk_mode"] == "data_parallel")
        self.assertEqual(splitk["instantiation_level"], 0)
        self.assertEqual(splitk["runtime_defaults"], {})
        self.assertEqual(splitk["allowed_runtime_sweeps"], ["shape_id", "m", "n", "k"])
        self.assertEqual(data_parallel["benchmark_target"], "03_bmg_gemm_streamk")

    def test_generated_level0_candidates_allow_stage_count_auto(self):
        shapes = {
            "schema_version": profiler.SCHEMA_VERSION,
            "generated_at": profiler.now_iso(),
            "shape_set_id": "generated-level0",
            "source": "test",
            "shapes": [
                {
                    "shape_id": "rcr_bf16_128_128_32",
                    "layout": "rcr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 128,
                    "n": 128,
                    "k": 32,
                }
            ],
        }

        candidate_space = profiler.generate_candidate_space(
            shapes,
            profiler.default_constraints(),
            profiler.default_compiler_profiles(),
            allowed_runners=("benchmark",),
            catalog_source="generator",
            generator_arch="bmg",
            generator_instantiation_level=0,
        )

        self.assertEqual(candidate_space["kernel_catalog"]["catalog_source"], "generator")
        self.assertEqual(candidate_space["kernel_catalog"]["generator_instantiation_level"], 0)
        self.assertGreater(len(candidate_space["candidates"]), 0)
        self.assertTrue(all(candidate["stages"] == 0 for candidate in candidate_space["candidates"]))
        self.assertTrue(
            any(
                candidate["kernel_id"]
                == "cutlass3x_xe20_tensorop_gemm_bf16_bf16_f32_f32_bf16_dbf16_128x128x32_1x1x1_0_tnt_align8"
                for candidate in candidate_space["candidates"]
            )
        )

    def test_kernel_catalog_prefers_repo_json(self):
        catalog = profiler.load_persisted_kernel_catalog()

        self.assertEqual(catalog["catalog_version"], "level0-seed-catalog")
        kernel_ids = {entry["kernel_id"] for entry in catalog["kernels"]}
        self.assertIn("BmgGemmFP16FP16FP32_RCR_6", kernel_ids)

    def test_generated_seed_catalog_matches_persisted_catalog(self):
        generated = profiler.generated_level0_kernel_catalog()
        persisted = profiler.load_persisted_kernel_catalog()

        normalize = lambda catalog: sorted(catalog["kernels"], key=lambda entry: entry["kernel_id"])
        self.assertEqual(generated["catalog_version"], persisted["catalog_version"])
        self.assertEqual(generated["search_runtime_schema"], persisted["search_runtime_schema"])
        self.assertEqual(normalize(generated), normalize(persisted))

    def test_build_kernel_catalog_can_bridge_generator_space(self):
        catalog = profiler.build_kernel_catalog(
            dtypes=["bf16"],
            allowed_runners=("benchmark",),
            catalog_source="generator",
            generator_arch="bmg",
            generator_instantiation_level=1,
        )

        self.assertEqual(catalog["catalog_source"], "generator")
        self.assertEqual(catalog["generator_arch"], "bmg")
        self.assertEqual(catalog["generator_instantiation_level"], 1)
        self.assertEqual(catalog["catalog_version"], "generator-bmg-level1")
        self.assertGreater(len(catalog["kernels"]), 8)
        self.assertTrue(any(entry["streamk_mode"] == "streamk" for entry in catalog["kernels"]))
        self.assertTrue(any(entry["stages"] == 0 for entry in catalog["kernels"]))
        self.assertTrue(all(entry["benchmark_target"] == "cutlass_benchmarks_gemm_sycl" for entry in catalog["kernels"]))

    def test_xe_generator_emits_auto_stage_count_for_staged_operations(self):
        repo_root = Path(__file__).resolve().parents[3]
        gemm_operation_path = repo_root / "python" / "cutlass_library" / "gemm_operation.py"
        text = gemm_operation_path.read_text(encoding="utf-8")
        xe_auto = text.index("if operation.is_xe:")
        positive_stage = text.index("elif operation.tile_description.stages > 0:")
        self.assertLess(xe_auto, positive_stage)

    def test_generate_candidate_space_can_use_generator_catalog(self):
        shapes = profiler.default_shapes("bf16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()

        candidate_space = profiler.generate_candidate_space(
            shapes,
            constraints,
            profiles,
            allowed_runners=("benchmark",),
            catalog_source="generator",
            generator_arch="bmg",
            generator_instantiation_level=1,
        )

        self.assertEqual(candidate_space["kernel_catalog"]["catalog_source"], "generator")
        self.assertEqual(candidate_space["kernel_catalog"]["generator_instantiation_level"], 1)
        self.assertGreater(candidate_space["kernel_catalog"]["kernel_count"], 8)
        self.assertEqual(candidate_space["candidate_coverage"]["requested_layouts"], ["rcr"])
        self.assertEqual(candidate_space["candidate_coverage"]["accepted_candidate_count"], len(candidate_space["candidates"]))
        self.assertEqual(
            candidate_space["candidate_coverage"]["exception_count"],
            len(candidate_space["candidate_exceptions"]),
        )
        self.assertFalse(any(candidate["streamk_mode"] == "streamk" for candidate in candidate_space["candidates"]))
        self.assertEqual(
            candidate_space["candidate_exception_summary"][0]["reason"],
            "intel_xe_generated_streamk_tile_scheduler_unsupported",
        )
        self.assertEqual(
            candidate_space["candidate_exception_summary"][0]["count"],
            candidate_space["candidate_coverage"]["exception_count"],
        )
        self.assertTrue(candidate_space["candidate_exception_summary"][0]["sample_kernel_names"][0].endswith("_stream_k"))
        self.assertTrue(
            any(
                item["reason"] == "intel_xe_generated_streamk_tile_scheduler_unsupported"
                and item["layout"] == "rcr"
                and item["dtype_a"] == "bf16"
                for item in candidate_space["candidate_exceptions"]
            )
        )
        self.assertTrue(all(candidate["dtype_c"] == "f32" for candidate in candidate_space["candidates"]))
        self.assertTrue(all(candidate["dtype_acc"] == "f32" for candidate in candidate_space["candidates"]))
        self.assertTrue(all(candidate["source"] == "generator_manifest" for candidate in candidate_space["candidates"]))

    def test_write_config_routes_generated_kernels_through_library_runner(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            config_path = Path(tmpdir) / "screening.in"
            entry = {
                "bm_name": "generated__shape__screening__0",
                "stage": "screening",
                "attempt_index": 0,
                "shape": {
                    "shape_id": "shape",
                    "layout": "rcr",
                    "dtype_a": "f16",
                    "dtype_b": "f16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 16,
                    "n": 32,
                    "k": 64,
                },
                "candidate": {
                    "candidate_id": "generated",
                    "compiler_profile_id": "default",
                    "kernel_name": "cutlass3x_xe20_tensorop_gemm_f16_f16_f32_f32_f16_df16_128x128x32_1x1x1_2_tnt_align8",
                    "split_k": 1,
                    "source": "generator_manifest",
                    "runner": "benchmark",
                },
            }

            metadata = profiler.write_config([entry], config_path)

            line = config_path.read_text(encoding="utf-8").strip()
            self.assertTrue(line.startswith("cutlass_library_gemm "))
            self.assertIn("--operation_name=cutlass3x_xe20_tensorop_gemm_f16_f16_f32_f32_f16_df16_128x128x32_1x1x1_2_tnt_align8", line)
            self.assertIn("--layout=rcr", line)
            self.assertIn("--dtype_a=f16", line)
            self.assertIn("--dtype_c=f32", line)
            self.assertEqual(metadata["generated__shape__screening__0"]["kernel_name"], entry["candidate"]["kernel_name"])

    def test_library_benchmark_runner_recognizes_bf16_d_generated_kernels(self):
        repo_root = Path(__file__).resolve().parents[3]
        runner_path = repo_root / "benchmarks" / "gemm" / "benchmark_runner.hpp"
        text = runner_path.read_text(encoding="utf-8")

        self.assertIn('operation_name.find("_bf16_dbf16_")', text)
        self.assertIn(
            "run_typed<cutlass::bfloat16_t, cutlass::bfloat16_t, cutlass::bfloat16_t>",
            text,
        )

    def test_build_candidate_build_manifest_splits_compile_and_runtime_fields(self):
        shapes = profiler.default_shapes("bf16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles)

        manifest = profiler.build_candidate_build_manifest(candidate_space)

        self.assertEqual(manifest["search_runtime_schema"]["microbench_guided_defaults"]["grf_mode"], 256)
        self.assertEqual(len(manifest["variants"]), 8)
        self.assertEqual(manifest["selected_kernel_count"], 8)
        self.assertEqual(len(manifest["selected_kernel_list"]), 8)
        self.assertTrue(all(line.startswith("^") and line.endswith("$") for line in manifest["kernel_filter_file"]["lines"]))
        self.assertEqual(manifest["kernel_filter_file"]["recommended_cmake_var"], "KERNEL_FILTER_FILE")
        self.assertEqual(manifest["cmake_config"]["cmake_vars"]["CUTLASS_LIBRARY_OPERATIONS"], "gemm")
        self.assertEqual(manifest["cmake_config"]["cmake_vars"]["BENCHMARK_ENABLE_TESTING"], "OFF")
        self.assertEqual(manifest["cmake_config"]["cmake_vars"]["BENCHMARK_ENABLE_GTEST_TESTS"], "OFF")
        variant = manifest["variants"][0]
        self.assertIn("compile_time_variant", variant)
        self.assertIn("runtime_sweep", variant)
        self.assertEqual(variant["runtime_sweep"]["allowed_fields"], ["shape_id", "m", "n", "k"])

    def test_build_candidate_build_manifest_can_emit_kernel_batches(self):
        shapes = profiler.default_shapes("bf16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles)

        manifest = profiler.build_candidate_build_manifest(candidate_space, selected_kernel_batch_size=3)

        self.assertEqual(manifest["selected_kernel_count"], 8)
        self.assertEqual(manifest["selected_kernel_batch_size"], 3)
        self.assertEqual([batch["kernel_count"] for batch in manifest["selected_kernel_batches"]], [3, 3, 2])
        self.assertEqual(manifest["selected_kernel_batches"][0]["batch_id"], "selected_kernel_batch_000")
        self.assertEqual(manifest["selected_kernel_batches"][0]["selected_kernel_list"], manifest["selected_kernel_list"][:3])
        self.assertTrue(
            all(
                line.startswith("^") and line.endswith("$")
                for batch in manifest["selected_kernel_batches"]
                for line in batch["kernel_filter_file"]["lines"]
            )
        )

    def test_default_constraints_use_calibrated_slm_limit(self):
        constraints = profiler.default_constraints()

        self.assertEqual(constraints["limits"]["max_slm_kb"], 64)

    def test_default_compiler_profiles_split_build_and_runtime_config(self):
        profiles = profiler.default_compiler_profiles()

        self.assertEqual(
            profiles["build_config"]["selected_compile_variant"],
            "perf_default",
        )
        self.assertEqual(
            profiles["build_config"]["cmake_vars"]["CUTLASS_SYCL_PROFILING_ENABLED"],
            "OFF",
        )
        self.assertEqual(
            profiles["build_config"]["cmake_vars"]["CUTLASS_ENABLE_EXAMPLES"],
            "OFF",
        )
        self.assertEqual(
            profiles["build_config"]["cmake_vars"]["CUTLASS_ENABLE_TESTS"],
            "OFF",
        )
        self.assertNotIn("IGC_VISAOptions", profiles["build_config"]["compile_env"])
        self.assertEqual(
            profiles["runtime_config"]["runtime_env"]["ONEAPI_DEVICE_SELECTOR"],
            "level_zero:gpu",
        )
        self.assertEqual(profiles["profiles"][0]["runtime_env_override"], {})
        self.assertEqual(
            profiles["build_config"]["compile_env_variant_metadata"]["perf_default"]["status"],
            "validated",
        )

    def test_build_config_marks_128grf_as_experimental(self):
        profiles = profiler.default_compiler_profiles()

        experiment = profiles["build_config"]["compile_env_variants"]["perf_128grf_experiment"]
        metadata = profiles["build_config"]["compile_env_variant_metadata"]["perf_128grf_experiment"]

        self.assertEqual(experiment["IGC_TotalGRFNum"], "128")
        self.assertNotIn("SYCL_PROGRAM_COMPILE_OPTIONS", experiment)
        self.assertEqual(metadata["status"], "needs_validation")
        self.assertIn("Advisory only", metadata["notes"])
        self.assertIn("do not pass -cl-intel-256-GRF-per-thread", metadata["notes"])
        self.assertIn("do not use in production", metadata["notes"])

    def test_load_persisted_build_config_fallback_matches_experimental_variants(self):
        persisted = profiler.load_persisted_build_config(profiler.DEFAULT_BUILD_CONFIG_PATH)
        with tempfile.TemporaryDirectory() as tmpdir:
            fallback_path = Path(tmpdir) / "missing_build_config.json"
            build_config = profiler.load_persisted_build_config(fallback_path)

        variants = build_config["compile_env_variants"]

        self.assertIn("perf_default", variants)
        self.assertIn("perf_perfmodel", variants)
        self.assertIn("perf_128grf_experiment", variants)
        self.assertIn("perf_enableBCR", variants)
        self.assertIn("debug_with_lines", variants)
        self.assertEqual(
            variants["perf_perfmodel"]["IGC_VISAOptions"],
            "-perfmodel",
        )
        self.assertEqual(
            variants["perf_enableBCR"]["IGC_VISAOptions"],
            "-enableBCR",
        )
        self.assertEqual(
            variants["debug_with_lines"]["SYCL_PROGRAM_COMPILE_OPTIONS"],
            "-ze-opt-large-register-file -gline-tables-only",
        )
        self.assertEqual(
            variants["perf_default"]["SYCL_PROGRAM_COMPILE_OPTIONS"],
            "-ze-opt-large-register-file",
        )
        self.assertEqual(
            build_config["compile_env_variant_metadata"]["perf_128grf_experiment"]["status"],
            "needs_validation",
        )
        self.assertEqual(build_config["compile_env_variants"], persisted["compile_env_variants"])
        self.assertEqual(
            build_config["compile_env_variant_metadata"],
            persisted["compile_env_variant_metadata"],
        )

    def test_selected_runtime_env_ignores_compile_time_flags(self):
        profiles = profiler.default_compiler_profiles()

        runtime_env = profiler.selected_runtime_env(profiles, profiles["profiles"][0])

        self.assertEqual(runtime_env["ONEAPI_DEVICE_SELECTOR"], "level_zero:gpu")
        self.assertNotIn("IGC_ExtraOCLOptions", runtime_env)
        self.assertNotIn("IGC_VISAOptions", runtime_env)
        self.assertNotIn("SYCL_PROGRAM_COMPILE_OPTIONS", runtime_env)

    def test_resolve_hw_reference_spec_uses_calibrated_b60_data(self):
        spec = profiler.resolve_hw_reference_spec("bmg")

        self.assertEqual(spec["device_id"], "bmg_g21")
        self.assertEqual(spec["clock_mhz"], 2400)
        self.assertEqual(spec["peak_bf16_tflops"], 97.66)
        self.assertEqual(spec["measured_read_bw_gbps"], 538)
        self.assertEqual(spec["slm_per_xe_core_kb"], 64)
        self.assertEqual(spec["calibration_status"], "measured")
        self.assertEqual(spec["concurrent_sgs_per_xe_core_128grf"], 16)
        self.assertEqual(spec["grf_bytes_per_thread_128grf"], 4096)

    def test_resolve_hw_reference_spec_marks_b70_as_not_measured(self):
        spec = profiler.resolve_hw_reference_spec(hw_spec_id="bmg_g31")

        self.assertEqual(spec["device_id"], "bmg_g31")
        self.assertEqual(spec["calibration_status"], "not_measured")

    def test_select_compiler_profile_skips_failed_probe_profile(self):
        profiles = profiler.default_compiler_profiles()
        profiles["profiles"][0]["probe_status"] = "fail"

        selected = profiler.select_compiler_profile_id(profiles, tile_m=8, sg_count=4)

        self.assertNotEqual(selected, "bmg.small_tile.default")

    def test_parse_benchmark_log_maps_generated_bm_name(self):
        metadata = {
            "rcr_bf16bf16f32_tm8_tn128_tk32_sg1x4_st2_sk1__rcr_bf16_1_4096_14336__screening__0": {
                "shape_id": "rcr_bf16_1_4096_14336",
                "candidate_id": "rcr_bf16bf16f32_tm8_tn128_tk32_sg1x4_st2_sk1",
                "compiler_profile_id": "bmg.small_tile.default",
                "stage": "screening",
                "attempt_index": 0,
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 1,
                "n": 4096,
                "k": 14336,
                "kernel_name": "BmgGemmBF16BF16FP32_RCR_5",
            }
        }
        line = (
            "BmgGemmBF16BF16FP32_RCR_5/"
            "rcr_bf16bf16f32_tm8_tn128_tk32_sg1x4_st2_sk1__rcr_bf16_1_4096_14336__screening__0/"
            "1x4096x14336x1/manual_time avg_runtime_ms=0.412 best_runtime_ms=0.398 "
            "worst_runtime_ms=0.437 avg_tflops=1.13 avg_throughput=287.4\n"
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            log_path = Path(tmpdir) / "screening.log"
            log_path.write_text(line, encoding="utf-8")
            rows = profiler.parse_benchmark_log(log_path, metadata, run_id="screening")

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["status"], "pass")
        self.assertEqual(row["verify_status"], "pass")
        self.assertEqual(row["shape_id"], "rcr_bf16_1_4096_14336")
        self.assertEqual(row["avg_tflops"], "1.13")

    def test_parse_benchmark_log_does_not_treat_max_error_metric_as_failure(self):
        metadata = {
            "bm_case": {
                "shape_id": "shape_a",
                "candidate_id": "cand_a",
                "compiler_profile_id": "bmg.small_tile.default",
                "stage": "screening",
                "attempt_index": 0,
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 1,
                "n": 64,
                "k": 32,
            }
        }
        line = (
            "Kernel/bm_case/1x64x32x1/manual_time avg_runtime_ms=0.1 best_runtime_ms=0.09 "
            "worst_runtime_ms=0.11 avg_tflops=1.2 avg_throughput=3.4 max_error=0.0001\n"
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            log_path = Path(tmpdir) / "parser.log"
            log_path.write_text(line, encoding="utf-8")
            rows = profiler.parse_benchmark_log(log_path, metadata, run_id="screening")

        self.assertEqual(rows[0]["status"], "pass")
        self.assertEqual(rows[0]["failure_reason"], "")

    def test_parse_benchmark_log_detects_real_error_line(self):
        metadata = {
            "bm_case": {
                "shape_id": "shape_a",
                "candidate_id": "cand_a",
                "compiler_profile_id": "bmg.small_tile.default",
                "stage": "screening",
                "attempt_index": 0,
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 1,
                "n": 64,
                "k": 32,
            }
        }
        line = "Kernel/bm_case/1x64x32x1 ERROR OCCURRED can_implement failed\n"
        with tempfile.TemporaryDirectory() as tmpdir:
            log_path = Path(tmpdir) / "parser_error.log"
            log_path.write_text(line, encoding="utf-8")
            rows = profiler.parse_benchmark_log(log_path, metadata, run_id="screening")

        self.assertEqual(rows[0]["status"], "fail")
        self.assertIn("ERROR OCCURRED", rows[0]["failure_reason"])

    def test_parse_benchmark_log_reports_missing_generated_registry_entry(self):
        metadata = {
            "bm_case": {
                "shape_id": "shape_a",
                "candidate_id": "generated_kernel_candidate",
                "compiler_profile_id": "bmg.large_tile.default",
                "stage": "screening",
                "attempt_index": 0,
                "layout": "rcr",
                "dtype_a": "f16",
                "dtype_b": "f16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 1,
                "n": 64,
                "k": 32,
            }
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            log_path = Path(tmpdir) / "missing_registry.log"
            log_path.write_text(
                "terminate called after throwing an instance of 'std::runtime_error'\n"
                "  what():  Benchmark not found\n",
                encoding="utf-8",
            )
            rows = profiler.parse_benchmark_log(log_path, metadata, run_id="screening")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["status"], "fail")
        self.assertEqual(rows[0]["candidate_id"], "generated_kernel_candidate")
        self.assertIn("benchmark registry entry not found", rows[0]["failure_reason"])

    def test_dispatch_table_uses_confirmation_median(self):
        shapes = {
            "schema_version": profiler.SCHEMA_VERSION,
            "generated_at": profiler.now_iso(),
            "shape_set_id": "test",
            "source": "predefined",
            "shapes": [
                {
                    "shape_id": "shape_a",
                    "layout": "rcr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 1,
                    "n": 4096,
                    "k": 4096,
                }
            ],
        }
        rows = [
            {
                "run_id": "screening",
                "stage": "screening",
                "attempt_index": 0,
                "shape_id": "shape_a",
                "candidate_id": "cand_fast",
                "compiler_profile_id": "bmg.small_tile.default",
                "status": "pass",
                "verify_status": "pass",
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 1,
                "n": 4096,
                "k": 4096,
                "avg_runtime_ms": "0.4",
                "best_runtime_ms": "0.39",
                "worst_runtime_ms": "0.42",
                "avg_tflops": "1.4",
                "avg_throughput": "100",
                "max_error": "",
                "close_call_group": "",
                "failure_reason": "",
                "stdout_log": "a.log",
            },
            {
                "run_id": "confirm",
                "stage": "confirm",
                "attempt_index": 0,
                "shape_id": "shape_a",
                "candidate_id": "cand_fast",
                "compiler_profile_id": "bmg.small_tile.default",
                "status": "pass",
                "verify_status": "pass",
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 1,
                "n": 4096,
                "k": 4096,
                "avg_runtime_ms": "0.5",
                "best_runtime_ms": "0.49",
                "worst_runtime_ms": "0.51",
                "avg_tflops": "1.0",
                "avg_throughput": "90",
                "max_error": "",
                "close_call_group": "",
                "failure_reason": "",
                "stdout_log": "b.log",
            },
            {
                "run_id": "confirm",
                "stage": "confirm",
                "attempt_index": 1,
                "shape_id": "shape_a",
                "candidate_id": "cand_fast",
                "compiler_profile_id": "bmg.small_tile.default",
                "status": "pass",
                "verify_status": "pass",
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 1,
                "n": 4096,
                "k": 4096,
                "avg_runtime_ms": "0.45",
                "best_runtime_ms": "0.44",
                "worst_runtime_ms": "0.46",
                "avg_tflops": "1.1",
                "avg_throughput": "95",
                "max_error": "",
                "close_call_group": "",
                "failure_reason": "",
                "stdout_log": "c.log",
            },
            {
                "run_id": "confirm",
                "stage": "confirm",
                "attempt_index": 0,
                "shape_id": "shape_a",
                "candidate_id": "cand_runner_up",
                "compiler_profile_id": "bmg.small_tile.default",
                "status": "pass",
                "verify_status": "pass",
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 1,
                "n": 4096,
                "k": 4096,
                "avg_runtime_ms": "0.46",
                "best_runtime_ms": "0.45",
                "worst_runtime_ms": "0.47",
                "avg_tflops": "1.05",
                "avg_throughput": "94",
                "max_error": "",
                "close_call_group": "",
                "failure_reason": "",
                "stdout_log": "d.log",
            },
        ]

        dispatch = profiler.build_dispatch_table(rows, shapes, top_k=3, confirm_runs=2, close_call_threshold=10.0)

        self.assertEqual(len(dispatch["entries"]), 1)
        entry = dispatch["entries"][0]
        self.assertEqual(entry["candidate_id"], "cand_fast")
        self.assertTrue(entry["close_call"])
        self.assertAlmostEqual(entry["evidence"]["confirm_median_tflops"], 1.05)
        self.assertEqual(entry["evidence"]["selection_stage"], "confirm")
        self.assertEqual(entry["evidence"]["screening_rank"], 1)
        self.assertEqual(entry["evidence"]["confirm_samples"], 2)
        self.assertTrue(entry["evidence"]["confirm_complete"])
        self.assertEqual(dispatch["selection_summary"]["entries_with_confirmation"], 1)
        self.assertEqual(dispatch["selection_summary"]["close_calls"], 1)

    def test_runtime_dispatch_lookup_matches_exact_shape(self):
        dispatch_table = {
            "schema_version": profiler.SCHEMA_VERSION,
            "entries": [
                {
                    "shape_key": {
                        "layout": "rcr",
                        "dtype_a": "bf16",
                        "dtype_b": "bf16",
                        "dtype_c": "f32",
                        "dtype_acc": "f32",
                        "m": 128,
                        "n": 256,
                        "k": 64,
                    },
                    "candidate_id": "winner",
                    "compiler_profile_id": "bmg.default",
                }
            ],
        }

        result = profiler.lookup_dispatch_entry(
            dispatch_table,
            {
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": "128",
                "n": "256",
                "k": "64",
            },
        )

        self.assertEqual(result["status"], "found")
        self.assertEqual(result["match"], "exact")
        self.assertFalse(result["fallback"]["used"])
        self.assertEqual(result["entry"]["candidate_id"], "winner")

    def test_runtime_dispatch_lookup_reports_explicit_fallback(self):
        dispatch_table = {"schema_version": profiler.SCHEMA_VERSION, "entries": []}

        result = profiler.lookup_dispatch_entry(
            dispatch_table,
            {
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 128,
                "n": 256,
                "k": 64,
            },
            fallback_candidate_id="safe_default",
        )

        self.assertEqual(result["status"], "fallback")
        self.assertEqual(result["match"], "none")
        self.assertIsNone(result["entry"])
        self.assertTrue(result["fallback"]["used"])
        self.assertEqual(result["fallback"]["reason"], "shape_not_found")
        self.assertEqual(result["fallback"]["candidate_id"], "safe_default")

    def test_runtime_dispatch_loader_rejects_bad_schema_and_duplicates(self):
        with self.assertRaisesRegex(ValueError, "unsupported dispatch table schema_version"):
            profiler.validate_dispatch_table({"schema_version": "0.9", "entries": []})

        entry = {
            "shape_key": {
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 128,
                "n": 256,
                "k": 64,
            },
            "candidate_id": "winner",
        }
        with self.assertRaisesRegex(ValueError, "duplicate dispatch shape_key"):
            profiler.validate_dispatch_table(
                {"schema_version": profiler.SCHEMA_VERSION, "entries": [entry, dict(entry)]}
            )

    def test_runtime_dispatch_loader_accepts_file_path(self):
        dispatch_table = {
            "schema_version": profiler.SCHEMA_VERSION,
            "entries": [
                {
                    "shape_key": {
                        "layout": "rcr",
                        "dtype_a": "f16",
                        "dtype_b": "f16",
                        "dtype_c": "f32",
                        "dtype_acc": "f32",
                        "m": 64,
                        "n": 64,
                        "k": 32,
                    },
                    "candidate_id": "file_winner",
                }
            ],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            table_path = Path(tmpdir) / "optimal_dispatch_table.json"
            profiler.write_json(table_path, dispatch_table)

            result = profiler.lookup_dispatch_entry(
                table_path,
                {
                    "layout": "rcr",
                    "dtype_a": "f16",
                    "dtype_b": "f16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 64,
                    "n": 64,
                    "k": 32,
                },
            )

        self.assertEqual(result["status"], "found")
        self.assertEqual(result["entry"]["candidate_id"], "file_winner")

    def test_runtime_dispatch_cli_lookup_returns_json(self):
        dispatch_table = {
            "schema_version": profiler.SCHEMA_VERSION,
            "entries": [
                {
                    "shape_key": {
                        "layout": "rcr",
                        "dtype_a": "bf16",
                        "dtype_b": "bf16",
                        "dtype_c": "f32",
                        "dtype_acc": "f32",
                        "m": 128,
                        "n": 128,
                        "k": 32,
                    },
                    "candidate_id": "cli_winner",
                }
            ],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            table_path = Path(tmpdir) / "optimal_dispatch_table.json"
            profiler.write_json(table_path, dispatch_table)
            repo_root = Path(__file__).resolve().parents[3]
            script_path = repo_root / "test" / "benchmarks" / "intel_gemm_profiler.py"
            completed = subprocess.run(
                [
                    sys.executable,
                    str(script_path),
                    "--lookup-dispatch-table",
                    str(table_path),
                    "--lookup-layout",
                    "rcr",
                    "--lookup-dtype-a",
                    "bf16",
                    "--lookup-dtype-b",
                    "bf16",
                    "--lookup-dtype-c",
                    "f32",
                    "--lookup-dtype-acc",
                    "f32",
                    "--lookup-m",
                    "128",
                    "--lookup-n",
                    "128",
                    "--lookup-k",
                    "32",
                ],
                check=True,
                text=True,
                capture_output=True,
            )

        result = json.loads(completed.stdout)
        self.assertEqual(result["status"], "found")
        self.assertEqual(result["entry"]["candidate_id"], "cli_winner")

    def test_runtime_dispatch_cli_lookup_returns_fallback_json(self):
        dispatch_table = {"schema_version": profiler.SCHEMA_VERSION, "entries": []}
        with tempfile.TemporaryDirectory() as tmpdir:
            table_path = Path(tmpdir) / "optimal_dispatch_table.json"
            profiler.write_json(table_path, dispatch_table)
            repo_root = Path(__file__).resolve().parents[3]
            script_path = repo_root / "test" / "benchmarks" / "intel_gemm_profiler.py"
            completed = subprocess.run(
                [
                    sys.executable,
                    str(script_path),
                    "--lookup-dispatch-table",
                    str(table_path),
                    "--lookup-m",
                    "128",
                    "--lookup-n",
                    "128",
                    "--lookup-k",
                    "32",
                    "--fallback-candidate-id",
                    "safe_default",
                ],
                check=True,
                text=True,
                capture_output=True,
            )

        result = json.loads(completed.stdout)
        self.assertEqual(result["status"], "fallback")
        self.assertEqual(result["fallback"]["candidate_id"], "safe_default")

    def test_product_bundle_validation_reports_missing_required_artifacts(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            dispatch_path = tmp / "optimal_dispatch_table.json"
            profiler.write_json(dispatch_path, {"schema_version": profiler.SCHEMA_VERSION, "entries": []})
            bundle_path = tmp / "gemm_product_bundle_manifest.json"
            profiler.write_json(
                bundle_path,
                {
                    "schema_version": profiler.SCHEMA_VERSION,
                    "required_artifacts": [
                        {
                            "name": "optimal_dispatch_table",
                            "path": str(dispatch_path),
                            "required": True,
                            "exists": True,
                            "purpose": "dispatch",
                        },
                        {
                            "name": "gemm_profile_results",
                            "path": str(tmp / "missing.csv"),
                            "required": True,
                            "exists": False,
                            "purpose": "profile rows",
                        },
                    ],
                    "optional_artifacts": [],
                    "runtime_lookup": {
                        "dispatch_table": str(dispatch_path),
                        "key_fields": list(profiler.DISPATCH_KEY_FIELDS),
                        "cli_args_template": ["python3", "test/benchmarks/intel_gemm_profiler.py", "--lookup-dispatch-table", str(dispatch_path)],
                    },
                },
            )

            validation = profiler.validate_product_bundle_manifest(bundle_path)
            self.assertEqual(validation["status"], "fail")
            self.assertEqual(validation["missing_required_artifacts"], ["gemm_profile_results"])

            repo_root = Path(__file__).resolve().parents[3]
            script_path = repo_root / "test" / "benchmarks" / "intel_gemm_profiler.py"
            completed = subprocess.run(
                [sys.executable, str(script_path), "--validate-product-bundle", str(bundle_path)],
                text=True,
                capture_output=True,
            )

        self.assertNotEqual(completed.returncode, 0)
        cli_validation = json.loads(completed.stdout)
        self.assertEqual(cli_validation["status"], "fail")
        self.assertIn("gemm_profile_results", cli_validation["missing_required_artifacts"])

    def test_confirmation_median_can_override_screening_rank(self):
        shapes = {
            "schema_version": profiler.SCHEMA_VERSION,
            "generated_at": profiler.now_iso(),
            "shape_set_id": "test",
            "source": "predefined",
            "shapes": [
                {
                    "shape_id": "shape_a",
                    "layout": "rcr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 64,
                    "n": 4096,
                    "k": 4096,
                }
            ],
        }

        def row(stage, candidate_id, attempt_index, tflops, runtime):
            return {
                "run_id": stage,
                "stage": stage,
                "attempt_index": attempt_index,
                "shape_id": "shape_a",
                "candidate_id": candidate_id,
                "compiler_profile_id": "bmg.default",
                "status": "pass",
                "verify_status": "pass",
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 64,
                "n": 4096,
                "k": 4096,
                "avg_runtime_ms": str(runtime),
                "best_runtime_ms": str(runtime),
                "worst_runtime_ms": str(runtime),
                "avg_tflops": str(tflops),
                "avg_throughput": "",
                "max_error": "",
                "close_call_group": "",
                "failure_reason": "",
                "stdout_log": f"{candidate_id}.log",
            }

        rows = [
            row("screening", "screening_winner", 0, 10.0, 0.40),
            row("screening", "confirmed_winner", 0, 9.0, 0.45),
            row("confirm", "screening_winner", 0, 7.5, 0.54),
            row("confirm", "screening_winner", 1, 7.7, 0.52),
            row("confirm", "screening_winner", 2, 7.6, 0.53),
            row("confirm", "confirmed_winner", 0, 8.3, 0.49),
            row("confirm", "confirmed_winner", 1, 8.5, 0.47),
            row("confirm", "confirmed_winner", 2, 8.4, 0.48),
        ]

        dispatch = profiler.build_dispatch_table(
            rows,
            shapes,
            top_k=2,
            confirm_runs=3,
            close_call_threshold=5.0,
        )

        entry = dispatch["entries"][0]
        evidence = entry["evidence"]
        self.assertEqual(entry["candidate_id"], "confirmed_winner")
        self.assertEqual(evidence["selection_stage"], "confirm")
        self.assertEqual(evidence["screening_rank"], 2)
        self.assertEqual(evidence["confirm_samples"], 3)
        self.assertEqual(evidence["expected_confirm_samples"], 3)
        self.assertTrue(evidence["confirm_complete"])
        self.assertEqual(evidence["runner_up_screening_rank"], 1)
        self.assertAlmostEqual(evidence["confirm_median_tflops"], 8.4)
        self.assertAlmostEqual(evidence["runner_up_median_tflops"], 7.6)
        self.assertGreater(evidence["confirm_tflops_stdev"], 0.0)
        self.assertEqual(len(evidence["ranked_candidates"]), 2)
        self.assertEqual(dispatch["selection_summary"]["incomplete_confirmation_entries"], 0)

    def test_splitk_candidates_are_gated_to_large_shapes(self):
        shapes = profiler.default_shapes("bf16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(
            shapes, constraints, profiles, allowed_runners=("benchmark", "streamk_example")
        )

        small_shape = next(shape for shape in shapes["shapes"] if shape["shape_id"] == "rcr_bf16_8_4096_4096")
        large_shape = next(shape for shape in shapes["shapes"] if shape["shape_id"] == "rcr_bf16_1_4096_14336")

        small_candidates = profiler.choose_candidates_for_shape(small_shape, candidate_space["candidates"])
        large_candidates = profiler.choose_candidates_for_shape(large_shape, candidate_space["candidates"])

        self.assertFalse(any(candidate["split_k"] > 1 for candidate in small_candidates))
        self.assertTrue(any(candidate["split_k"] > 1 for candidate in large_candidates))

    def test_choose_candidates_handles_m16_and_m128_boundaries(self):
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        shapes = {
            "schema_version": profiler.SCHEMA_VERSION,
            "generated_at": profiler.now_iso(),
            "shape_set_id": "boundary",
            "source": "test",
            "shapes": [
                {
                    "shape_id": "shape_m16",
                    "layout": "rcr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 16,
                    "n": 4096,
                    "k": 4096,
                },
                {
                    "shape_id": "shape_m128",
                    "layout": "rcr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 128,
                    "n": 4096,
                    "k": 8192,
                },
            ],
        }
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles)

        m16_candidates = profiler.choose_candidates_for_shape(shapes["shapes"][0], candidate_space["candidates"])
        m128_candidates = profiler.choose_candidates_for_shape(shapes["shapes"][1], candidate_space["candidates"])

        self.assertTrue(any(candidate["tile_m"] == 128 for candidate in m16_candidates))
        self.assertTrue(any(candidate["tile_m"] == 256 for candidate in m128_candidates))

    def test_f16_streamk_modes_use_streamk_runner(self):
        shapes = profiler.default_shapes("f16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(
            shapes, constraints, profiles, allowed_runners=("benchmark", "streamk_example")
        )

        streamk = next(candidate for candidate in candidate_space["candidates"] if candidate["streamk_mode"] == "streamk")
        data_parallel = next(candidate for candidate in candidate_space["candidates"] if candidate["streamk_mode"] == "data_parallel")
        splitk = next(candidate for candidate in candidate_space["candidates"] if candidate["split_k"] == 2)
        self.assertEqual(streamk["dtype_a"], "f16")
        self.assertEqual(streamk["runner"], "streamk_example")
        self.assertEqual(data_parallel["dtype_a"], "f16")
        self.assertEqual(data_parallel["runner"], "streamk_example")
        self.assertEqual(splitk["dtype_a"], "f16")
        self.assertEqual(splitk["runner"], "streamk_example")

    def test_f16_candidate_space_includes_large_tile(self):
        shapes = profiler.default_shapes("f16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles)

        candidate_ids = {candidate["candidate_id"] for candidate in candidate_space["candidates"]}
        self.assertIn("rcr_f16f16f32_tm256_tn256_tk32_sg8x4_st2_sk1", candidate_ids)
        self.assertIn("rcr_f16f16f32_tm128_tn256_tk32_sg4x4_st2_sk1", candidate_ids)

    def test_dry_run_shapes_use_tiny_shape_set(self):
        shapes = profiler.dry_run_shapes("bf16")

        self.assertEqual(shapes["source"], "dry_run")
        self.assertEqual(len(shapes["shapes"]), 1)
        shape = shapes["shapes"][0]
        self.assertEqual((shape["m"], shape["n"], shape["k"]), (1, 64, 32))
        self.assertEqual(shape["runtime_defaults"], {})

    def test_build_dpas_probe_entry_uses_small_benchmark_candidate(self):
        shapes = profiler.default_shapes("bf16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles, allowed_runners=("benchmark",))

        entry = profiler.build_dpas_probe_entry(shapes, candidate_space)

        self.assertIsNotNone(entry)
        self.assertEqual(entry["stage"], "dpas_probe")
        self.assertEqual(entry["candidate"]["candidate_id"], "rcr_bf16bf16f32_tm8_tn64_tk32_sg1x4_st2_sk1")
        self.assertEqual(entry["shape"]["shape_id"], "rcr_bf16_8_4096_4096")

    def test_build_phase_a_probe_entries_use_dynamic_shapes(self):
        shapes = {
            "schema_version": profiler.SCHEMA_VERSION,
            "generated_at": profiler.now_iso(),
            "shape_set_id": "custom",
            "source": "custom",
            "shapes": [
                {
                    "shape_id": "custom_small",
                    "layout": "rcr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 4,
                    "n": 2048,
                    "k": 2048,
                },
                {
                    "shape_id": "custom_mid",
                    "layout": "rcr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 48,
                    "n": 4096,
                    "k": 4096,
                },
                {
                    "shape_id": "custom_big",
                    "layout": "rcr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 192,
                    "n": 4096,
                    "k": 8192,
                },
            ],
        }
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles, allowed_runners=("benchmark",))

        entries = profiler.build_phase_a_probe_entries(shapes, candidate_space)
        shape_ids = {entry["shape"]["shape_id"] for entry in entries}

        self.assertIn("custom_small", shape_ids)
        self.assertIn("custom_mid", shape_ids)
        self.assertIn("custom_big", shape_ids)

    def test_build_compiler_profile_probe_entries_matches_candidate_classes(self):
        shapes = profiler.default_shapes("bf16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles, allowed_runners=("benchmark",))

        entries = profiler.build_compiler_profile_probe_entries(shapes, candidate_space, profiles)

        self.assertEqual(len(entries), 3)
        probe_ids = {entry["compiler_profile_probe_id"] for entry in entries}
        self.assertIn("bmg.small_tile.default", probe_ids)
        self.assertIn("bmg.medium_tile.default", probe_ids)
        self.assertIn("bmg.large_tile.default", probe_ids)
        self.assertEqual(entries[0]["compiler_profile_id"], entries[0]["compiler_profile_probe_id"])

    def test_static_probe_constraints_record_splitk_feedback(self):
        constraints = profiler.default_constraints()
        env_caps = {
            "executables": {
                "benchmark_available": True,
                "streamk_example_available": False,
            }
        }

        updated = profiler.apply_static_probe_constraints(constraints, env_caps)

        self.assertEqual(updated["constraint_source"], "phase_a_static_probe")
        self.assertEqual(updated["limits"]["max_split_k"], 1)
        self.assertEqual(updated["allowed_values"]["split_k"], [1])
        self.assertEqual(updated["probe_feedback"]["mode"], "static")
        self.assertEqual(updated["probe_feedback"]["actions"][0]["action"], "limit_split_k")
        self.assertEqual(updated["probe_feedback"]["actions"][0]["reason"], "streamk_example_unavailable")

    def test_run_probe_constraints_record_pruning_feedback(self):
        constraints = profiler.default_constraints()
        static_constraints = profiler.apply_static_probe_constraints(
            constraints,
            {
                "executables": {
                    "benchmark_available": True,
                    "streamk_example_available": True,
                }
            },
        )
        failing_candidate = "rcr_bf16bf16f32_tm64_tn128_tk32_sg4x4_st2_sk1"
        anomalous_candidate = "rcr_bf16bf16f32_tm128_tn128_tk32_sg4x4_st2_sk1"
        probe_rows = [
            {
                "candidate_id": failing_candidate,
                "shape_id": "shape_a",
                "status": "fail",
                "failure_reason": "Disposition Failed",
                "split_k": "1",
                "stdout_log": "probe.log",
            },
            {
                "candidate_id": "rcr_bf16bf16f32_tm8_tn64_tk32_sg1x4_st2_sk1",
                "shape_id": "shape_b",
                "status": "pass",
                "avg_tflops": "10.0",
                "split_k": "1",
                "stdout_log": "probe.log",
            },
        ]
        anomaly_report = {
            "anomalies": [{"candidate_id": anomalous_candidate}],
            "auto_block_rules": [
                {
                    "rule_id": f"probe.auto_block.anomaly.{anomalous_candidate}",
                    "match": {
                        "tile_m": 128,
                        "tile_n": 128,
                        "tile_k": 32,
                        "sg_m": 4,
                        "sg_n": 4,
                        "split_k": 1,
                    },
                    "reason": "large_tile_slower_than_small_tile",
                    "evidence_tflops": 1.0,
                }
            ],
        }

        updated = profiler.apply_run_probe_constraints(static_constraints, probe_rows, anomaly_report)
        feedback = updated["probe_feedback"]

        self.assertEqual(updated["constraint_source"], "phase_a_run_probe")
        self.assertEqual(updated["limits"]["max_split_k"], 1)
        self.assertEqual(feedback["mode"], "run")
        self.assertEqual(feedback["probe_rows"], 2)
        self.assertEqual(feedback["passed_probe_rows"], 1)
        self.assertEqual(feedback["failed_probe_rows"], 1)
        self.assertEqual(feedback["anomaly_count"], 1)
        self.assertEqual(feedback["auto_block_rule_count"], 1)
        self.assertEqual(feedback["blocked_rule_count"], 2)
        action_reasons = {action["reason"] for action in feedback["actions"]}
        self.assertIn("no_successful_split_k_probe", action_reasons)
        self.assertIn("probe_failure", action_reasons)
        self.assertIn("large_tile_slower_than_small_tile", action_reasons)
        failure_rule = next(rule for rule in updated["blocked_rules"] if rule["rule_id"] == f"probe.blocked.{failing_candidate}")
        self.assertEqual(failure_rule["source"], "phase_a_probe_failure")
        self.assertEqual(failure_rule["evidence"]["stdout_log"], "probe.log")
        self.assertTrue(
            profiler.blocked(
                {
                    "tile_m": 128,
                    "tile_n": 128,
                    "tile_k": 32,
                    "sg_m": 4,
                    "sg_n": 4,
                    "stages": 2,
                    "split_k": 1,
                    "grf_mode": 256,
                },
                updated,
            )
        )

    def test_dry_run_workflow_uses_minimal_shape_set_and_no_confirmation(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            args = profiler.build_parser().parse_args(
                [
                    "--workspace",
                    tmpdir,
                    "--dtype",
                    "f16",
                    "--dry-run",
                    "--skip-run",
                ]
            )
            outputs = profiler.workflow(args)

            self.assertTrue(outputs["dry_run"])
            shapes_doc = profiler.read_json(Path(tmpdir) / "inputs" / "gemm_target_shapes.json")
            self.assertEqual(shapes_doc["source"], "dry_run")
            phase_a_summary = profiler.read_json(Path(tmpdir) / "reports" / "phase_a_summary.json")
            self.assertEqual(phase_a_summary["probe_mode"], "dry_run_off")
            self.assertEqual(phase_a_summary["probe_feedback"]["mode"], "default")
            phase_b_summary = profiler.read_json(Path(tmpdir) / "reports" / "phase_b_summary.json")
            self.assertEqual(phase_b_summary["candidate_count"], 8)
            bundle = profiler.read_json(Path(outputs["artifact_bundle_manifest"]))
            self.assertEqual(bundle["schema_version"], profiler.SCHEMA_VERSION)
            self.assertEqual(bundle["workspace"], str(Path(tmpdir).resolve()))
            self.assertEqual(bundle["missing_required_artifacts"], [])
            required_names = {artifact["name"] for artifact in bundle["required_artifacts"]}
            self.assertIn("optimal_dispatch_table", required_names)
            self.assertIn("gemm_profile_results", required_names)
            self.assertIn("phase_b_summary", required_names)
            optimal_artifact = next(
                artifact for artifact in bundle["required_artifacts"] if artifact["name"] == "optimal_dispatch_table"
            )
            self.assertGreater(optimal_artifact["size_bytes"], 0)
            self.assertRegex(optimal_artifact["sha256"], r"^[0-9a-f]{64}$")
            self.assertIn("reference_comparison", bundle["missing_optional_artifacts"])
            self.assertEqual(
                bundle["runtime_lookup"]["key_fields"],
                ["layout", "dtype_a", "dtype_b", "dtype_c", "dtype_acc", "m", "n", "k"],
            )
            self.assertIn("--lookup-dispatch-table", bundle["runtime_lookup"]["cli_template"])
            self.assertIn(outputs["optimal_dispatch_table"], bundle["runtime_lookup"]["cli_args_template"])
            validation = profiler.validate_product_bundle_manifest(outputs["artifact_bundle_manifest"])
            self.assertEqual(validation["status"], "pass")
            self.assertEqual(validation["missing_required_artifacts"], [])
            self.assertEqual(validation["integrity_errors"], [])
            self.assertEqual(validation["dispatch_entry_count"], 0)

            repo_root = Path(__file__).resolve().parents[3]
            script_path = repo_root / "test" / "benchmarks" / "intel_gemm_profiler.py"
            completed = subprocess.run(
                [
                    sys.executable,
                    str(script_path),
                    "--validate-product-bundle",
                    outputs["artifact_bundle_manifest"],
                ],
                check=True,
                text=True,
                capture_output=True,
            )
            cli_validation = json.loads(completed.stdout)
            self.assertEqual(cli_validation["status"], "pass")

            Path(outputs["results_csv"]).write_text("tampered\n", encoding="utf-8")
            tampered_validation = profiler.validate_product_bundle_manifest(outputs["artifact_bundle_manifest"])
            self.assertEqual(tampered_validation["status"], "fail")
            self.assertTrue(
                any("gemm_profile_results" in error for error in tampered_validation["integrity_errors"])
            )

    def test_skip_run_workflow_can_emit_generator_backed_catalog(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            args = profiler.build_parser().parse_args(
                [
                    "--workspace",
                    tmpdir,
                    "--dtype",
                    "bf16",
                    "--skip-run",
                    "--kernel-catalog-source",
                    "generator",
                    "--generator-arch",
                    "bmg",
                    "--generator-instantiation-level",
                    "1",
                    "--googlebenchmark-dir",
                    str(Path(tmpdir) / "googlebenchmark-src"),
                    "--cmake-cxx-compiler",
                    "icpx",
                    "--candidate-build-batch-size",
                    "4",
                ]
            )
            outputs = profiler.workflow(args)

            kernel_catalog = profiler.read_json(Path(outputs["kernel_catalog"]))
            candidate_space = profiler.read_json(Path(outputs["candidate_space"]))
            build_manifest = profiler.read_json(Path(outputs["build_manifest"]))
            cmake_config = profiler.read_json(Path(outputs["candidate_build_cmake_config"]))
            build_plan = profiler.read_json(Path(outputs["candidate_build_plan"]))
            build_summary = profiler.read_json(Path(outputs["candidate_build_summary"]))
            preflight_summary = profiler.read_json(Path(outputs["candidate_build_preflight_summary"]))
            selected_kernel_list = Path(outputs["selected_kernel_list"]).read_text(encoding="utf-8").splitlines()
            selected_kernel_filter = Path(outputs["selected_kernel_filter"]).read_text(encoding="utf-8").splitlines()

            self.assertEqual(kernel_catalog["catalog_source"], "generator")
            self.assertEqual(kernel_catalog["generator_instantiation_level"], 1)
            self.assertEqual(candidate_space["kernel_catalog"]["catalog_source"], "generator")
            self.assertTrue(any(entry["stages"] == 0 for entry in kernel_catalog["kernels"]))
            self.assertEqual(candidate_space["candidate_coverage"]["accepted_candidate_count"], len(candidate_space["candidates"]))
            self.assertEqual(candidate_space["candidate_coverage"]["exception_count"], len(candidate_space["candidate_exceptions"]))
            self.assertTrue(candidate_space["candidate_exception_summary"][0]["sample_kernel_names"][0].endswith("_stream_k"))
            self.assertGreater(len(build_manifest["variants"]), 8)
            self.assertFalse(any("_stream_k" in variant["kernel_id"] for variant in build_manifest["variants"]))
            self.assertTrue(
                any(
                    item["reason"] == "intel_xe_generated_streamk_tile_scheduler_unsupported"
                    for item in candidate_space["candidate_exceptions"]
                )
            )
            self.assertEqual(build_manifest["selected_kernel_count"], len(selected_kernel_list))
            self.assertEqual(build_manifest["selected_kernel_list"], selected_kernel_list)
            self.assertEqual(build_manifest["kernel_filter_file"]["lines"], selected_kernel_filter)
            self.assertEqual(build_manifest["selected_kernel_batch_size"], 4)
            self.assertGreater(len(build_manifest["selected_kernel_batches"]), 1)
            first_batch = build_manifest["selected_kernel_batches"][0]
            first_batch_filter = Path(first_batch["kernel_filter_path"]).read_text(encoding="utf-8").splitlines()
            self.assertEqual(first_batch_filter, first_batch["kernel_filter_file"]["lines"])
            self.assertEqual(cmake_config["kernel_filter_cmake_var"], "KERNEL_FILTER_FILE")
            self.assertEqual(cmake_config["cmake_vars"]["CUTLASS_LIBRARY_INSTANTIATION_LEVEL"], "1")
            self.assertEqual(build_plan["build_target"], "cutlass_benchmarks_gemm_sycl")
            self.assertTrue(build_plan["benchmark_exe"].endswith("/benchmarks/gemm/cutlass_benchmarks_gemm_sycl"))
            self.assertEqual(build_plan["kernel_filter_file"], outputs["selected_kernel_filter"])
            self.assertEqual(build_plan["selected_kernel_batch_size"], 4)
            self.assertEqual(build_plan["selected_kernel_batches"][0]["kernel_filter_path"], first_batch["kernel_filter_path"])
            self.assertEqual(len(build_plan["batch_preflight_plans"]), len(build_manifest["selected_kernel_batches"]))
            first_preflight = build_plan["batch_preflight_plans"][0]
            self.assertEqual(first_preflight["batch_id"], first_batch["batch_id"])
            self.assertEqual(first_preflight["kernel_filter_file"], first_batch["kernel_filter_path"])
            self.assertIn("candidate_batch_preflight/selected_kernel_batch_000", first_preflight["build_dir"])
            self.assertIn(f"-DKERNEL_FILTER_FILE={first_batch['kernel_filter_path']}", first_preflight["configure_command"])
            self.assertTrue(first_preflight["benchmark_exe"].endswith("/benchmarks/gemm/cutlass_benchmarks_gemm_sycl"))
            self.assertEqual(build_plan["cmake_vars"]["KERNEL_FILTER_FILE"], outputs["selected_kernel_filter"])
            self.assertEqual(build_plan["googlebenchmark_dir"], str(Path(tmpdir) / "googlebenchmark-src"))
            self.assertEqual(build_plan["cmake_vars"]["GOOGLEBENCHMARK_DIR"], str(Path(tmpdir) / "googlebenchmark-src"))
            self.assertIn(f"-DGOOGLEBENCHMARK_DIR={Path(tmpdir) / 'googlebenchmark-src'}", build_plan["configure_command"])
            self.assertEqual(build_plan["cmake_cxx_compiler"], "icpx")
            self.assertEqual(build_plan["cmake_vars"]["CMAKE_CXX_COMPILER"], "icpx")
            self.assertIn("-DCMAKE_CXX_COMPILER=icpx", build_plan["configure_command"])
            self.assertIn("-DCUTLASS_LIBRARY_INSTANTIATION_LEVEL=1", build_plan["configure_command"])
            self.assertIn(f"-DKERNEL_FILTER_FILE={outputs['selected_kernel_filter']}", build_plan["configure_command"])
            self.assertEqual(build_plan["build_command"][4], "cutlass_benchmarks_gemm_sycl")
            self.assertEqual(build_summary["status"], "not_run")
            self.assertEqual(preflight_summary["status"], "not_run")

    def test_execute_candidate_build_plan_runs_configure_and_build_steps(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            build_plan = {
                "schema_version": profiler.SCHEMA_VERSION,
                "generated_at": profiler.now_iso(),
                "build_target": "fake_benchmark",
                "benchmark_exe": str(tmp / "build" / "benchmarks" / "gemm" / "fake_benchmark"),
                "configure_command": ["python3", "-c", "print('configure ok')"],
                "build_command": ["python3", "-c", "print('build ok')"],
            }

            summary = profiler.execute_candidate_build_plan(build_plan, tmp / "logs")

        self.assertEqual(summary["status"], "pass")
        self.assertEqual(summary["benchmark_exe"], build_plan["benchmark_exe"])
        self.assertEqual([step["step"] for step in summary["steps"]], ["configure", "build"])
        self.assertTrue(all(step["status"] == "pass" for step in summary["steps"]))

    def test_execute_candidate_build_plan_returns_failure_summary(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            build_plan = {
                "schema_version": profiler.SCHEMA_VERSION,
                "generated_at": profiler.now_iso(),
                "build_target": "fake_benchmark",
                "benchmark_exe": str(tmp / "build" / "benchmarks" / "gemm" / "fake_benchmark"),
                "selected_kernel_count": 3,
                "kernel_filter_file": str(tmp / "selected_kernel_filter.list"),
                "configure_command": ["python3", "-c", "print('configure ok')"],
                "build_command": ["python3", "-c", "import sys; print('build failed'); sys.exit(7)"],
            }

            summary = profiler.execute_candidate_build_plan(build_plan, tmp / "logs")
            build_log_text = Path(summary["steps"][1]["log"]).read_text(encoding="utf-8")

        self.assertEqual(summary["status"], "fail")
        self.assertEqual(summary["failure_step"], "build")
        self.assertIn("Candidate benchmark build failed", summary["failure_reason"])
        self.assertEqual(summary["selected_kernel_count"], 3)
        self.assertEqual(summary["kernel_filter_file"], build_plan["kernel_filter_file"])
        self.assertEqual([step["step"] for step in summary["steps"]], ["configure", "build"])
        self.assertEqual(summary["steps"][1]["returncode"], 7)
        self.assertTrue(build_log_text.strip().endswith("build failed"))

    def test_execute_candidate_build_preflight_plans_reports_batch_status(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            build_plan = {
                "schema_version": profiler.SCHEMA_VERSION,
                "generated_at": profiler.now_iso(),
                "batch_preflight_plans": [
                    {
                        "schema_version": profiler.SCHEMA_VERSION,
                        "generated_at": profiler.now_iso(),
                        "build_target": "fake_benchmark",
                        "batch_id": "selected_kernel_batch_000",
                        "batch_index": 0,
                        "kernel_count": 2,
                        "benchmark_exe": str(tmp / "batch0" / "fake_benchmark"),
                        "kernel_filter_file": str(tmp / "selected_kernel_filter_part000.list"),
                        "configure_command": ["python3", "-c", "print('configure batch0')"],
                        "build_command": ["python3", "-c", "print('build batch0')"],
                    },
                    {
                        "schema_version": profiler.SCHEMA_VERSION,
                        "generated_at": profiler.now_iso(),
                        "build_target": "fake_benchmark",
                        "batch_id": "selected_kernel_batch_001",
                        "batch_index": 1,
                        "kernel_count": 1,
                        "benchmark_exe": str(tmp / "batch1" / "fake_benchmark"),
                        "kernel_filter_file": str(tmp / "selected_kernel_filter_part001.list"),
                        "configure_command": ["python3", "-c", "print('configure batch1')"],
                        "build_command": ["python3", "-c", "import sys; print('build batch1 failed'); sys.exit(9)"],
                    },
                ],
            }

            summary = profiler.execute_candidate_build_preflight_plans(build_plan, tmp / "logs")
            failed_log_text = Path(summary["batches"][1]["steps"][1]["log"]).read_text(encoding="utf-8")

        self.assertEqual(summary["status"], "fail")
        self.assertEqual(summary["batch_count"], 2)
        self.assertEqual(summary["passed_batches"], 1)
        self.assertEqual(summary["failed_batches"], 1)
        self.assertEqual(summary["batches"][0]["status"], "pass")
        self.assertEqual(summary["batches"][1]["status"], "fail")
        self.assertEqual(summary["batches"][1]["batch_id"], "selected_kernel_batch_001")
        self.assertEqual(summary["batches"][1]["kernel_count"], 1)
        self.assertIn("candidate_build_preflight_selected_kernel_batch_001.log", summary["batches"][1]["steps"][1]["log"])
        self.assertTrue(failed_log_text.strip().endswith("build batch1 failed"))

    def test_workflow_persists_candidate_build_failure_summary(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            fake_cmake = tmp / "fake_cmake.py"
            fake_cmake.write_text(
                "#!/usr/bin/env python3\n"
                "import sys\n"
                "print('fake cmake ' + ' '.join(sys.argv[1:]))\n"
                "sys.exit(5 if '--build' in sys.argv else 0)\n",
                encoding="utf-8",
            )
            fake_cmake.chmod(0o755)
            args = profiler.build_parser().parse_args(
                [
                    "--workspace",
                    str(tmp / "workspace"),
                    "--dtype",
                    "bf16",
                    "--skip-run",
                    "--probe-mode",
                    "off",
                    "--kernel-catalog-source",
                    "generator",
                    "--build-candidate-benchmark",
                    "--cmake-source-dir",
                    str(tmp),
                    "--benchmark-build-dir",
                    str(tmp / "build"),
                    "--cmake-cxx-compiler",
                    "icpx",
                ]
            )
            original_build_candidate_build_plan = profiler.build_candidate_build_plan

            def fake_build_candidate_build_plan(*args, **kwargs):
                plan = original_build_candidate_build_plan(*args, **kwargs)
                plan["configure_command"][0] = str(fake_cmake)
                plan["build_command"][0] = str(fake_cmake)
                return plan

            workflow_globals = profiler.workflow.__globals__
            original_workflow_build_candidate_build_plan = workflow_globals["build_candidate_build_plan"]
            profiler.build_candidate_build_plan = fake_build_candidate_build_plan
            workflow_globals["build_candidate_build_plan"] = fake_build_candidate_build_plan
            try:
                with self.assertRaisesRegex(RuntimeError, "Candidate benchmark build failed"):
                    profiler.workflow(args)
            finally:
                profiler.build_candidate_build_plan = original_build_candidate_build_plan
                workflow_globals["build_candidate_build_plan"] = original_workflow_build_candidate_build_plan

            summary_path = tmp / "workspace" / "reports" / "candidate_build_summary.json"
            summary = profiler.read_json(summary_path)
            build_log_exists = Path(summary["steps"][1]["log"]).exists()

        self.assertEqual(summary["status"], "fail")
        self.assertEqual(summary["failure_step"], "build")
        self.assertEqual(summary["steps"][0]["status"], "pass")
        self.assertEqual(summary["steps"][1]["status"], "fail")
        self.assertEqual(summary["steps"][1]["returncode"], 5)
        self.assertTrue(build_log_exists)

    def test_workflow_persists_candidate_build_preflight_failure_summary(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            args = profiler.build_parser().parse_args(
                [
                    "--workspace",
                    str(tmp / "workspace"),
                    "--dtype",
                    "bf16",
                    "--dry-run",
                    "--probe-mode",
                    "off",
                    "--kernel-catalog-source",
                    "generator",
                    "--candidate-build-batch-size",
                    "1",
                    "--run-candidate-build-preflight",
                    "--cmake-source-dir",
                    str(tmp),
                    "--benchmark-build-dir",
                    str(tmp / "build"),
                ]
            )
            fake_log = tmp / "fake_preflight_build.log"

            def fake_execute_candidate_build_preflight_plans(build_plan, log_dir, shell_init="", timeout=None):
                fake_log.write_text("preflight failed\n", encoding="utf-8")
                batches = []
                preflight_plans = build_plan["batch_preflight_plans"] or [
                    {
                        "build_target": "fake_benchmark",
                        "benchmark_exe": str(tmp / "fake_benchmark"),
                        "batch_id": "selected_kernel_batch_000",
                        "kernel_count": 1,
                    }
                ]
                for index, plan in enumerate(preflight_plans):
                    batches.append(
                        {
                            "schema_version": profiler.SCHEMA_VERSION,
                            "generated_at": profiler.now_iso(),
                            "status": "fail",
                            "failure_step": "build",
                            "failure_reason": f"Candidate benchmark build failed with status fail. See {fake_log}.",
                            "build_target": plan["build_target"],
                            "benchmark_exe": plan["benchmark_exe"],
                            "batch_id": plan["batch_id"],
                            "batch_index": index,
                            "kernel_count": plan["kernel_count"],
                            "steps": [
                                {"step": "configure", "status": "pass", "returncode": 0, "command": "configure", "log": str(fake_log)},
                                {"step": "build", "status": "fail", "returncode": 6, "command": "build", "log": str(fake_log)},
                            ],
                        }
                    )
                return {
                    "schema_version": profiler.SCHEMA_VERSION,
                    "generated_at": profiler.now_iso(),
                    "status": "fail",
                    "batch_count": len(batches),
                    "passed_batches": 0,
                    "failed_batches": len(batches),
                    "failure_reason": batches[0]["failure_reason"],
                    "batches": batches,
                }

            workflow_globals = profiler.workflow.__globals__
            original_workflow_execute_preflight = workflow_globals["execute_candidate_build_preflight_plans"]
            profiler.execute_candidate_build_preflight_plans = fake_execute_candidate_build_preflight_plans
            workflow_globals["execute_candidate_build_preflight_plans"] = fake_execute_candidate_build_preflight_plans
            try:
                with self.assertRaisesRegex(RuntimeError, "Candidate benchmark build failed"):
                    profiler.workflow(args)
            finally:
                profiler.execute_candidate_build_preflight_plans = original_workflow_execute_preflight
                workflow_globals["execute_candidate_build_preflight_plans"] = original_workflow_execute_preflight

            summary_path = tmp / "workspace" / "reports" / "candidate_build_preflight_summary.json"
            summary = profiler.read_json(summary_path)
            build_log_exists = Path(summary["batches"][0]["steps"][1]["log"]).exists()

        self.assertEqual(summary["status"], "fail")
        self.assertGreaterEqual(summary["batch_count"], 1)
        self.assertEqual(summary["passed_batches"], 0)
        self.assertEqual(summary["failed_batches"], summary["batch_count"])
        self.assertEqual(summary["batches"][0]["status"], "fail")
        self.assertEqual(summary["batches"][0]["batch_id"], "selected_kernel_batch_000")
        self.assertEqual(summary["batches"][0]["steps"][1]["returncode"], 6)
        self.assertTrue(build_log_exists)

    def test_build_candidate_benchmark_requires_phase_a_inputs_or_probe_off(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            args = profiler.build_parser().parse_args(
                [
                    "--workspace",
                    tmpdir,
                    "--kernel-catalog-source",
                    "generator",
                    "--build-candidate-benchmark",
                    "--benchmark-exe",
                    str(Path(tmpdir) / "missing_benchmark"),
                ]
            )

            with self.assertRaisesRegex(ValueError, "--probe-mode=off"):
                profiler.workflow(args)

    def test_build_candidate_benchmark_allows_probe_off_without_prebuilt_benchmark(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            args = profiler.build_parser().parse_args(
                [
                    "--workspace",
                    tmpdir,
                    "--kernel-catalog-source",
                    "generator",
                    "--build-candidate-benchmark",
                    "--probe-mode",
                    "off",
                    "--benchmark-exe",
                    str(Path(tmpdir) / "missing_benchmark"),
                ]
            )

            profiler.validate_candidate_auto_build_mode(args, dry_run_mode=False, probe_mode=args.probe_mode)

    def test_workflow_can_limit_generator_candidates_to_compiled_kernel_list(self):
        shapes = profiler.dry_run_shapes("f16")
        candidate_space = profiler.generate_candidate_space(
            shapes,
            profiler.default_constraints(),
            profiler.default_compiler_profiles(),
            allowed_runners=("benchmark",),
            catalog_source="generator",
            generator_arch="bmg",
            generator_instantiation_level=1,
        )
        selected_kernel = candidate_space["candidates"][0]["kernel_id"]

        with tempfile.TemporaryDirectory() as tmpdir:
            compiled_kernel_list = Path(tmpdir) / "compiled_kernels.list"
            compiled_kernel_list.write_text(f"^{selected_kernel}$\n", encoding="utf-8")
            args = profiler.build_parser().parse_args(
                [
                    "--workspace",
                    tmpdir,
                    "--dtype",
                    "f16",
                    "--dry-run",
                    "--skip-run",
                    "--kernel-catalog-source",
                    "generator",
                    "--generator-arch",
                    "bmg",
                    "--generator-instantiation-level",
                    "1",
                    "--compiled-kernel-list",
                    str(compiled_kernel_list),
                ]
            )
            outputs = profiler.workflow(args)

            filtered_candidate_space = profiler.read_json(Path(outputs["candidate_space"]))
            build_manifest = profiler.read_json(Path(outputs["build_manifest"]))
            build_plan = profiler.read_json(Path(outputs["candidate_build_plan"]))

        self.assertEqual(filtered_candidate_space["compiled_kernel_filter"]["kernel_count"], 1)
        self.assertEqual(filtered_candidate_space["compiled_kernel_filter"]["matched_candidate_count"], 1)
        self.assertEqual(build_manifest["selected_kernel_list"], [selected_kernel])
        self.assertEqual(build_plan["selected_kernel_count"], 1)

    def test_dispatch_table_reports_low_efficiency_winner(self):
        shapes = {
            "schema_version": profiler.SCHEMA_VERSION,
            "generated_at": profiler.now_iso(),
            "shape_set_id": "efficiency",
            "source": "test",
            "shapes": [
                {
                    "shape_id": "shape_peak_gap",
                    "layout": "rcr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 256,
                    "n": 4096,
                    "k": 8192,
                }
            ],
        }
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles)
        rows = [
            {
                "run_id": "confirm",
                "stage": "confirm",
                "attempt_index": 0,
                "shape_id": "shape_peak_gap",
                "candidate_id": "rcr_bf16bf16f32_tm16_tn64_tk32_sg2x4_st2_sk1",
                "compiler_profile_id": "bmg.small_tile.default",
                "status": "pass",
                "verify_status": "pass",
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 256,
                "n": 4096,
                "k": 8192,
                "avg_runtime_ms": "0.57",
                "best_runtime_ms": "0.56",
                "worst_runtime_ms": "0.58",
                "avg_tflops": "10.0",
                "avg_throughput": "0",
                "max_error": "",
                "close_call_group": "",
                "failure_reason": "",
                "stdout_log": "winner.log",
            }
        ]

        dispatch = profiler.build_dispatch_table(
            rows,
            shapes,
            top_k=3,
            confirm_runs=1,
            close_call_threshold=3.0,
            candidate_space=candidate_space,
            hw_spec=profiler.resolve_hw_reference_spec("bmg"),
        )

        entry = dispatch["entries"][0]
        self.assertEqual(entry["candidate_id"], "rcr_bf16bf16f32_tm16_tn64_tk32_sg2x4_st2_sk1")
        self.assertEqual(entry["efficiency_warning"], "winner_efficiency_below_40pct_peak")
        self.assertLess(entry["selected_efficiency"], 0.4)

    def test_phase_b_summary_surfaces_low_efficiency_warnings(self):
        candidate_space = {"candidates": [{}], "kernel_catalog": {"catalog_version": "test-catalog"}}
        dispatch_table = {
            "entries": [
                {
                    "shape_id": "shape_peak_gap",
                    "candidate_id": "cand_a",
                    "selected_efficiency": 0.31,
                    "peak_tflops": 97.66,
                    "efficiency_warning": "winner_efficiency_below_40pct_peak",
                }
            ]
        }
        summary = {"rows": 10, "passed": 10, "failed": 0}

        phase_b_summary = profiler.build_phase_b_summary(candidate_space, dispatch_table, summary)

        self.assertEqual(len(phase_b_summary["low_efficiency_warnings"]), 1)
        self.assertEqual(
            phase_b_summary["low_efficiency_warnings"][0]["warning"],
            "winner_efficiency_below_40pct_peak",
        )

    def test_dispatch_table_skips_low_efficiency_warning_for_memory_bound_winner(self):
        shapes = {
            "schema_version": profiler.SCHEMA_VERSION,
            "generated_at": profiler.now_iso(),
            "shape_set_id": "memory-bound",
            "source": "test",
            "shapes": [
                {
                    "shape_id": "decode_shape",
                    "layout": "rcr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 1,
                    "n": 4096,
                    "k": 14336,
                }
            ],
        }
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles)
        rows = [
            {
                "run_id": "confirm",
                "stage": "confirm",
                "attempt_index": 0,
                "shape_id": "decode_shape",
                "candidate_id": "rcr_bf16bf16f32_tm8_tn128_tk32_sg1x8_st2_sk1",
                "compiler_profile_id": "bmg.small_tile.default",
                "status": "pass",
                "verify_status": "pass",
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 1,
                "n": 4096,
                "k": 14336,
                "avg_runtime_ms": "0.235",
                "best_runtime_ms": "0.230",
                "worst_runtime_ms": "0.240",
                "avg_tflops": "0.5",
                "avg_throughput": "0",
                "max_error": "",
                "close_call_group": "",
                "failure_reason": "",
                "stdout_log": "decode.log",
            }
        ]

        dispatch = profiler.build_dispatch_table(
            rows,
            shapes,
            top_k=1,
            confirm_runs=1,
            close_call_threshold=3.0,
            candidate_space=candidate_space,
            hw_spec=profiler.resolve_hw_reference_spec("bmg"),
        )

        entry = dispatch["entries"][0]
        self.assertLess(entry["selected_efficiency"], 0.4)
        self.assertEqual(entry["efficiency_warning"], "")

    def test_generate_candidate_space_rejects_unsupported_layout(self):
        shapes = {
            "schema_version": profiler.SCHEMA_VERSION,
            "generated_at": profiler.now_iso(),
            "shape_set_id": "unsupported-layout",
            "source": "test",
            "shapes": [
                {
                    "shape_id": "shape_ccc",
                    "layout": "ccc",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 64,
                    "n": 256,
                    "k": 256,
                }
            ],
        }

        with self.assertRaisesRegex(ValueError, "Unsupported layouts in shapes: ccc"):
            profiler.generate_candidate_space(shapes, profiler.default_constraints(), profiler.default_compiler_profiles())

    def test_generate_candidate_space_supports_registered_rrr_layout(self):
        shapes = {
            "schema_version": profiler.SCHEMA_VERSION,
            "generated_at": profiler.now_iso(),
            "shape_set_id": "rrr-layout",
            "source": "test",
            "shapes": [
                {
                    "shape_id": "rrr_bf16_512_256_32",
                    "layout": "rrr",
                    "dtype_a": "bf16",
                    "dtype_b": "bf16",
                    "dtype_c": "f32",
                    "dtype_acc": "f32",
                    "m": 512,
                    "n": 256,
                    "k": 32,
                }
            ],
        }

        candidate_space = profiler.generate_candidate_space(
            shapes,
            profiler.default_constraints(),
            profiler.default_compiler_profiles(),
            allowed_runners=("benchmark",),
        )

        self.assertEqual(len(candidate_space["candidates"]), 1)
        self.assertEqual(
            candidate_space["candidates"][0]["kernel_name"],
            "BmgGemmBF16BF16FP32_RRR_TileShape_512_256_32",
        )

    def test_phase_a_summary_includes_dpas_probe(self):
        summary = profiler.build_phase_a_summary(
            {
                "probe_mode": "run",
                "dpas_baseline_probe": {"status": "pass", "avg_tflops": "1.23"},
                "compiler_flags_probe": {"results": [{"compiler_profile_id": "bmg.small_tile.default", "status": "pass"}]},
            },
            profiler.default_constraints(),
            [],
        )

        self.assertEqual(summary["dpas_baseline_probe"]["status"], "pass")
        self.assertEqual(summary["compiler_flags_probe"]["results"][0]["compiler_profile_id"], "bmg.small_tile.default")

    def test_apply_probe_results_to_profiles_marks_selected_profiles(self):
        profiles = profiler.default_compiler_profiles()
        summary = {
            "results": [
                {"compiler_profile_id": "bmg.small_tile.default", "status": "pass", "avg_tflops": "1.0", "avg_runtime_ms": "0.1"},
                {"compiler_profile_id": "bmg.medium_tile.default", "status": "fail", "avg_tflops": "", "avg_runtime_ms": ""},
                {"compiler_profile_id": "bmg.large_tile.default", "status": "pass", "avg_tflops": "2.0", "avg_runtime_ms": "0.2"},
            ],
            "selected_profile_ids": {"small_tile": "bmg.small_tile.default", "large_tile": "bmg.large_tile.default"},
        }

        updated = profiler.apply_probe_results_to_profiles(profiles, summary)

        by_id = {profile["compiler_profile_id"]: profile for profile in updated["profiles"]}
        self.assertEqual(by_id["bmg.small_tile.default"]["probe_status"], "pass")
        self.assertTrue(by_id["bmg.small_tile.default"]["probe_selected"])
        self.assertEqual(by_id["bmg.medium_tile.default"]["probe_status"], "fail")
        self.assertFalse(by_id["bmg.medium_tile.default"]["probe_selected"])

    def test_build_compiler_flags_probe_summary_groups_by_candidate_class(self):
        profiles = profiler.default_compiler_profiles()
        rows = [
            {
                "compiler_profile_id": "bmg.small_tile.default",
                "candidate_id": "cand_small_a",
                "shape_id": "shape_small_a",
                "status": "pass",
                "avg_tflops": "1.0",
                "avg_runtime_ms": "0.10",
                "stdout_log": "small_a.log",
            },
            {
                "compiler_profile_id": "bmg.small_tile.default",
                "candidate_id": "cand_small_b",
                "shape_id": "shape_small_b",
                "status": "pass",
                "avg_tflops": "1.4",
                "avg_runtime_ms": "0.08",
                "stdout_log": "small_b.log",
            },
            {
                "compiler_profile_id": "bmg.large_tile.default",
                "candidate_id": "cand_large",
                "shape_id": "shape_large",
                "status": "fail",
                "avg_tflops": "",
                "avg_runtime_ms": "",
                "stdout_log": "large.log",
            },
        ]

        summary = profiler.build_compiler_flags_probe_summary(rows, profiles)

        self.assertEqual(summary["selected_profile_ids"]["small_tile"], "bmg.small_tile.default")
        results_by_id = {item["compiler_profile_id"]: item for item in summary["results"]}
        self.assertEqual(results_by_id["bmg.small_tile.default"]["candidate_class"], "small_tile")
        self.assertEqual(results_by_id["bmg.small_tile.default"]["samples"], 2)
        self.assertEqual(results_by_id["bmg.small_tile.default"]["avg_tflops"], "1.2")
        self.assertEqual(results_by_id["bmg.large_tile.default"]["candidate_class"], "large_tile")

    def test_build_compiler_flags_probe_summary_handles_empty_pass_metrics(self):
        profiles = profiler.default_compiler_profiles()
        rows = [
            {
                "compiler_profile_id": "bmg.small_tile.default",
                "candidate_id": "cand_small",
                "shape_id": "shape_small",
                "status": "pass",
                "avg_tflops": "",
                "avg_runtime_ms": "",
                "stdout_log": "small.log",
            }
        ]

        summary = profiler.build_compiler_flags_probe_summary(rows, profiles)

        result = summary["results"][0]
        self.assertEqual(result["avg_tflops"], "")
        self.assertEqual(result["avg_runtime_ms"], "")
        self.assertEqual(summary["selected_profile_ids"]["small_tile"], "bmg.small_tile.default")

    def test_runner_environment_metadata_uses_shared_schema_version(self):
        metadata = profiler.collect_environment_metadata("", "missing-benchmark", "missing-streamk")

        self.assertEqual(metadata["schema_version"], profiler.SCHEMA_VERSION)
        self.assertRegex(metadata["python_version"], r"^\d+\.\d+\.\d+")

    def test_run_benchmark_returns_timeout(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            log_path = Path(tmpdir) / "timeout.log"
            process, timed_out, reason = profiler.run_benchmark(
                ["python3", "-c", "import sys, time; sys.stdout.buffer.write(b'partial\\xff output\\n'); sys.stdout.flush(); time.sleep(2)"],
                log_path,
                timeout=1,
            )
            log_text = log_path.read_text(encoding="utf-8")

        self.assertTrue(timed_out)
        self.assertEqual(process.returncode, 124)
        self.assertIn("timeout after 1s", reason)
        self.assertIn("partial", log_text)
        self.assertIn("TIMEOUT: timeout after 1s", log_text)

    def test_run_entries_with_benchmark_supports_chunking(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            fake_exe = tmp / "fake_benchmark.py"
            fake_exe.write_text(
                "#!/usr/bin/env python3\n"
                "import sys\n"
                "config = sys.argv[1].split('=', 1)[1]\n"
                "for line in open(config, encoding='utf-8'):\n"
                "    bm = line.split('--bm_name=', 1)[1].split()[0]\n"
                "    print(f'BM/{bm}/manual_time avg_runtime_ms=1.0 avg_tflops=2.0')\n",
                encoding="utf-8",
            )
            fake_exe.chmod(0o755)
            shape = {
                "shape_id": "shape_a",
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 64,
                "n": 4096,
                "k": 4096,
            }
            candidate = {
                "candidate_id": "cand_a",
                "compiler_profile_id": "profile_a",
                "kernel_name": "kernel_a",
                "split_k": 1,
            }
            entries = [
                {"bm_name": f"bm_{idx}", "stage": "screening", "attempt_index": idx, "shape": shape, "candidate": candidate}
                for idx in range(5)
            ]

            rows, commands = profiler.run_entries_with_benchmark(
                entries,
                tmp / "screening.in",
                tmp / "screening_manifest.json",
                tmp / "screening.log",
                str(fake_exe),
                chunk_size=2,
            )

            self.assertEqual(len(rows), 5)
            self.assertEqual(len(commands), 3)
            self.assertTrue((tmp / "screening_part000.in").exists())
            self.assertTrue((tmp / "screening_part002.log").exists())

    def test_run_entries_with_batch_benchmarks_routes_by_kernel_id(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            fake_exe = tmp / "fake_benchmark.py"
            fake_exe.write_text(
                "#!/usr/bin/env python3\n"
                "import sys\n"
                "config = sys.argv[1].split('=', 1)[1]\n"
                "for line in open(config, encoding='utf-8'):\n"
                "    bm = line.split('--bm_name=', 1)[1].split()[0]\n"
                "    print(f'BM/{bm}/manual_time avg_runtime_ms=1.0 avg_tflops=2.0')\n",
                encoding="utf-8",
            )
            fake_exe.chmod(0o755)
            shape = {
                "shape_id": "shape_a",
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 64,
                "n": 4096,
                "k": 4096,
            }
            candidate_a = {
                "candidate_id": "cand_a",
                "compiler_profile_id": "profile_a",
                "kernel_name": "kernel_a",
                "kernel_id": "kernel_a",
                "split_k": 1,
            }
            candidate_b = {
                "candidate_id": "cand_b",
                "compiler_profile_id": "profile_a",
                "kernel_name": "kernel_b",
                "kernel_id": "kernel_b",
                "split_k": 1,
            }
            entries = [
                {"bm_name": "bm_a", "stage": "screening", "attempt_index": 0, "shape": shape, "candidate": candidate_a},
                {"bm_name": "bm_b", "stage": "screening", "attempt_index": 0, "shape": shape, "candidate": candidate_b},
            ]
            batch_plan_by_kernel_id = {
                "kernel_a": {"batch_id": "selected_kernel_batch_000", "benchmark_exe": str(fake_exe)},
                "kernel_b": {"batch_id": "selected_kernel_batch_001", "benchmark_exe": str(fake_exe)},
            }

            rows, commands, logs = profiler.run_entries_with_batch_benchmarks(
                entries,
                tmp / "screening.in",
                tmp / "screening_manifest.json",
                tmp / "screening.log",
                batch_plan_by_kernel_id,
            )

            self.assertEqual(len(rows), 2)
            self.assertEqual(len(commands), 2)
            self.assertEqual(len(logs), 2)
            self.assertTrue((tmp / "screening_selected_kernel_batch_000.in").exists())
            self.assertTrue((tmp / "screening_selected_kernel_batch_001.log").exists())
            self.assertTrue(any("selected_kernel_batch_000" in log for log in logs))
            self.assertTrue(any("selected_kernel_batch_001" in log for log in logs))

    def test_run_entries_with_streamk_example_supports_streamk_and_dp_modes(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            logs_dir = Path(tmpdir) / "logs"
            logs_dir.mkdir()
            fake_exe = Path(tmpdir) / "fake_streamk.py"
            fake_exe.write_text(
                "#!/usr/bin/env python3\n"
                "import sys\n"
                "print('ARGS:' + ' '.join(sys.argv[1:]))\n"
                "print('Cutlass GEMM Performance: [1.23]TFlop/s (0.45)ms')\n"
                "print('Disposition: Passed')\n",
                encoding='utf-8',
            )
            fake_exe.chmod(0o755)
            shape = {
                "shape_id": "shape_a",
                "layout": "rcr",
                "dtype_a": "bf16",
                "dtype_b": "bf16",
                "dtype_c": "f32",
                "dtype_acc": "f32",
                "m": 64,
                "n": 4096,
                "k": 4096,
            }
            streamk_candidate = {
                "candidate_id": "cand_streamk",
                "compiler_profile_id": "bmg.medium_tile.default",
                "dtype_a": "bf16",
                "kernel_name": "03_bmg_gemm_streamk_streamk_bf16",
                "split_k": 1,
                "streamk_mode": "streamk",
            }
            data_parallel_candidate = {
                "candidate_id": "cand_dp",
                "compiler_profile_id": "bmg.medium_tile.default",
                "dtype_a": "bf16",
                "kernel_name": "03_bmg_gemm_streamk_dp_bf16",
                "split_k": 1,
                "streamk_mode": "data_parallel",
            }
            entries = [
                {"bm_name": "cand_streamk__shape_a__screening__0", "stage": "screening", "attempt_index": 0, "shape": shape, "candidate": streamk_candidate},
                {"bm_name": "cand_dp__shape_a__screening__0", "stage": "screening", "attempt_index": 0, "shape": shape, "candidate": data_parallel_candidate},
            ]

            rows, commands = profiler.run_entries_with_streamk_example(entries, logs_dir, str(fake_exe))

        self.assertEqual(len(rows), 2)
        self.assertTrue(all(row["status"] == "pass" for row in rows))
        self.assertIn("--dtype=bf16", commands[0])
        self.assertNotIn("--splitk", commands[0])
        self.assertNotIn("--dp", commands[0])
        self.assertIn("--dp", commands[1])
        self.assertNotIn("--splitk", commands[1])

    def test_static_probe_disables_splitk_without_streamk_binary(self):
        constraints = profiler.default_constraints()
        env_caps = {
            "executables": {
                "benchmark_available": True,
                "streamk_example_available": False,
            }
        }
        probed = profiler.apply_static_probe_constraints(constraints, env_caps)
        self.assertEqual(probed["limits"]["max_split_k"], 1)
        self.assertEqual(probed["allowed_values"]["split_k"], [1])

    def test_run_probe_adds_blocked_rule_for_failed_candidate(self):
        constraints = profiler.default_constraints()
        rows = [
            {
                "candidate_id": "rcr_bf16bf16f32_tm8_tn64_tk32_sg1x4_st2_sk2",
                "status": "fail",
                "split_k": "2",
            }
        ]
        probed = profiler.apply_run_probe_constraints(constraints, rows)
        self.assertEqual(probed["limits"]["max_split_k"], 1)
        self.assertEqual(probed["allowed_values"]["split_k"], [1])
        self.assertEqual(len(probed["blocked_rules"]), 1)
        self.assertEqual(probed["blocked_rules"][0]["match"]["tile_m"], 8)

    def test_detect_probe_anomalies_blocks_large_tile_regression(self):
        shapes = profiler.default_shapes("bf16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles, allowed_runners=("benchmark",))
        hw_spec = profiler.resolve_hw_reference_spec("bmg")
        small_candidate = min(
            (candidate for candidate in candidate_space["candidates"] if candidate["candidate_class"] == "small_tile"),
            key=lambda candidate: (candidate["tile_m"], candidate["tile_n"]),
        )
        large_candidate = max(
            (candidate for candidate in candidate_space["candidates"] if candidate["candidate_class"] == "large_tile"),
            key=lambda candidate: (candidate["tile_m"], candidate["tile_n"]),
        )
        probe_rows = [
            {
                "candidate_id": small_candidate["candidate_id"],
                "shape_id": "rcr_bf16_8_4096_4096",
                "status": "pass",
                "avg_tflops": "2.997",
            },
            {
                "candidate_id": small_candidate["candidate_id"],
                "shape_id": "rcr_bf16_64_4096_4096",
                "status": "pass",
                "avg_tflops": "18.9",
            },
            {
                "candidate_id": large_candidate["candidate_id"],
                "shape_id": "rcr_bf16_256_4096_8192",
                "status": "pass",
                "avg_tflops": "2.282",
            },
        ]

        report = profiler.detect_probe_anomalies(probe_rows, shapes, candidate_space, hw_spec)

        self.assertEqual(report["hw_spec"], "bmg_g21")
        self.assertEqual(len(report["anomalies"]), 1)
        anomaly = report["anomalies"][0]
        self.assertEqual(anomaly["candidate_id"], large_candidate["candidate_id"])
        self.assertEqual(anomaly["spec_anomaly"], "severely_below_spec")
        self.assertTrue(anomaly["cross_anomaly"].startswith("large_tile_slower_than_"))
        self.assertEqual(report["auto_block_rules"][0]["rule_id"], f"probe.auto_block.anomaly.{large_candidate['candidate_id']}")

    def test_apply_run_probe_constraints_appends_anomaly_auto_block(self):
        constraints = profiler.default_constraints()
        rows = []
        anomaly_report = {
            "auto_block_rules": [
                {
                    "rule_id": "probe.auto_block.anomaly.test_candidate",
                    "match": {"tile_m": 256, "tile_n": 256, "tile_k": 32, "sg_m": 8, "sg_n": 4, "split_k": 1},
                    "reason": "severely_below_spec",
                    "evidence_tflops": 2.282,
                }
            ]
        }

        probed = profiler.apply_run_probe_constraints(constraints, rows, anomaly_report=anomaly_report)

        self.assertEqual(probed["limits"]["max_slm_kb"], 64)
        self.assertEqual(len(probed["blocked_rules"]), 1)
        self.assertEqual(probed["blocked_rules"][0]["rule_id"], "probe.auto_block.anomaly.test_candidate")

    def test_compute_efficiency_bounds_handles_memory_bound_decode_shape(self):
        hw_spec = profiler.resolve_hw_reference_spec("bmg")
        shape = profiler.default_shapes("bf16")["shapes"][0]
        candidate = {
            "tile_m": 8,
            "tile_n": 64,
            "tile_k": 32,
            "sg_m": 1,
            "sg_n": 4,
            "stages": 2,
            "dtype_a": "bf16",
        }

        min_eff, max_eff = profiler.compute_efficiency_bounds(shape, candidate, hw_spec)

        self.assertLess(min_eff, max_eff)
        self.assertLess(max_eff, 0.02)

    def test_detect_probe_anomalies_above_expected_is_report_only(self):
        shapes = profiler.default_shapes("bf16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles, allowed_runners=("benchmark",))
        hw_spec = profiler.resolve_hw_reference_spec("bmg")
        small_candidate = min(
            (candidate for candidate in candidate_space["candidates"] if candidate["candidate_class"] == "small_tile"),
            key=lambda candidate: (candidate["tile_m"], candidate["tile_n"]),
        )
        probe_rows = [
            {
                "candidate_id": small_candidate["candidate_id"],
                "shape_id": "rcr_bf16_8_4096_4096",
                "status": "pass",
                "avg_tflops": "20.0",
            }
        ]

        report = profiler.detect_probe_anomalies(probe_rows, shapes, candidate_space, hw_spec)

        self.assertEqual(report["hw_spec_calibration_status"], "measured")
        self.assertEqual(report["anomalies"][0]["spec_anomaly"], "above_expected")
        self.assertEqual(report["anomalies"][0]["auto_action"], "reported")
        self.assertEqual(report["auto_block_rules"], [])

    def test_detect_probe_anomalies_uses_fastest_cross_reference(self):
        shapes = profiler.default_shapes("bf16")
        constraints = profiler.default_constraints()
        profiles = profiler.default_compiler_profiles()
        candidate_space = profiler.generate_candidate_space(shapes, constraints, profiles, allowed_runners=("benchmark",))
        hw_spec = profiler.resolve_hw_reference_spec("bmg")
        small_candidates = sorted(
            (candidate for candidate in candidate_space["candidates"] if candidate["candidate_class"] == "small_tile"),
            key=lambda candidate: (candidate["tile_m"], candidate["tile_n"]),
        )
        large_candidate = max(
            (candidate for candidate in candidate_space["candidates"] if candidate["candidate_class"] == "large_tile"),
            key=lambda candidate: (candidate["tile_m"], candidate["tile_n"]),
        )
        probe_rows = [
            {
                "candidate_id": small_candidates[0]["candidate_id"],
                "shape_id": "rcr_bf16_8_4096_4096",
                "status": "pass",
                "avg_tflops": "10.0",
            },
            {
                "candidate_id": small_candidates[1]["candidate_id"],
                "shape_id": "rcr_bf16_64_4096_4096",
                "status": "pass",
                "avg_tflops": "20.0",
            },
            {
                "candidate_id": large_candidate["candidate_id"],
                "shape_id": "rcr_bf16_256_4096_8192",
                "status": "pass",
                "avg_tflops": "2.0",
            },
        ]

        report = profiler.detect_probe_anomalies(probe_rows, shapes, candidate_space, hw_spec)

        large_anomaly = next(
            item for item in report["anomalies"] if item["candidate_id"] == large_candidate["candidate_id"]
        )
        self.assertEqual(
            large_anomaly["cross_anomaly"],
            f"large_tile_slower_than_{small_candidates[1]['candidate_id']}",
        )

    def test_build_ali_gemm_docs_extracts_supported_shapes_and_reference(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ali.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "v1.6"
            sheet.append(["", "", "", "", "", "bf16", "", "", "", "", "", "", "", "", "", "f16", "", "", "", "int8", "", "", ""])
            sheet.append([
                "",
                "M",
                "N",
                "K",
                "Type",
                "oneMKL(BLAS)",
                "oneDNN(matmul)",
                "PyTorch->oneDNN",
                "Sycl_TLA(00_base)",
                "Sycl_TLA(00_padded)",
                "Sycl_TLA(00_sycl_q)",
                "Sycl_TLA(03_streamk)",
                "Sycl_TLA(03_dp)",
                "",
                "",
                "oneMKL(BLAS)",
                "oneDNN(matmul)",
                "PyTorch->oneDNN",
                "SYCL-TLA(XeTLA)",
                "oneMKL(BLAS)",
                "oneDNN(matmul)",
                "PyTorch->oneDNN",
                "SYCL-TLA(XeTLA)",
            ])
            sheet.append([
                "",
                8192,
                4096,
                4096,
                "Compute Bounded",
                None,
                158.0,
                155.0,
                155.8,
                155.9,
                155.8,
                156.7,
                160.2,
                None,
                None,
                None,
                119.6,
                157.2,
                153.3,
                150.1,
                315.0,
                312.8,
                None,
                300.0,
            ])
            workbook.save(workbook_path)

            shapes_doc, reference_doc = profiler.build_ali_gemm_docs(workbook_path)

        self.assertEqual(len(shapes_doc["shapes"]), 2)
        shape_ids = {shape["shape_id"] for shape in shapes_doc["shapes"]}
        self.assertIn("rcr_bf16_8192_4096_4096", shape_ids)
        self.assertIn("rcr_f16_8192_4096_4096", shape_ids)
        bf16_entry = next(entry for entry in reference_doc["entries"] if entry["dtype_a"] == "bf16")
        self.assertEqual(bf16_entry["reference_provider"], "Sycl_TLA(03_dp)")
        self.assertEqual(bf16_entry["reference_tflops"], 160.2)
        self.assertTrue(any(item["dtype"] == "int8" for item in reference_doc["skipped_entries"]))

    def test_workflow_can_derive_shapes_and_reference_from_ali_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ali.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "v1.6"
            sheet.append(["", "", "", "", "", "bf16", "", "", "", "", "", "", "", "", "", "f16", "", "", "", "int8", "", "", ""])
            sheet.append([
                "",
                "M",
                "N",
                "K",
                "Type",
                "oneMKL(BLAS)",
                "oneDNN(matmul)",
                "PyTorch->oneDNN",
                "Sycl_TLA(00_base)",
                "Sycl_TLA(00_padded)",
                "Sycl_TLA(00_sycl_q)",
                "Sycl_TLA(03_streamk)",
                "Sycl_TLA(03_dp)",
                "",
                "",
                "oneMKL(BLAS)",
                "oneDNN(matmul)",
                "PyTorch->oneDNN",
                "SYCL-TLA(XeTLA)",
                "oneMKL(BLAS)",
                "oneDNN(matmul)",
                "PyTorch->oneDNN",
                "SYCL-TLA(XeTLA)",
            ])
            sheet.append(["", 64, 4096, 4096, "Prefill", None, None, None, 100.0, 101.0, 99.0, 98.0, 102.0, None, None, None, None, None, None, None, None, None, None])
            workbook.save(workbook_path)

            args = profiler.build_parser().parse_args(
                [
                    "--workspace",
                    tmpdir,
                    "--ali-workbook",
                    str(workbook_path),
                    "--skip-run",
                    "--probe-mode",
                    "off",
                ]
            )
            outputs = profiler.workflow(args)

            shapes_doc = profiler.read_json(Path(outputs["workspace"]) / "inputs" / "gemm_target_shapes.json")
            reference_doc = profiler.read_json(Path(outputs["reference_doc"]))
            comparison = profiler.read_json(Path(outputs["reference_comparison"]))

        self.assertEqual(shapes_doc["source"], str(workbook_path))
        self.assertEqual(len(shapes_doc["shapes"]), 1)
        self.assertEqual(reference_doc["dataset_id"], "ali_gemm_perf_reference")
        self.assertEqual(comparison["summary"]["reference_entries"], 1)
        self.assertEqual(comparison["summary"]["missing_dispatch"], 1)

    def test_workflow_limits_ali_shapes_and_reference_entries(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ali.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "v1.6"
            sheet.append(["", "", "", "", "", "bf16", "", "", "", "", "", "", "", "", "", "f16", "", "", "", "int8", "", "", ""])
            sheet.append([
                "",
                "M",
                "N",
                "K",
                "Type",
                "oneMKL(BLAS)",
                "oneDNN(matmul)",
                "PyTorch->oneDNN",
                "Sycl_TLA(00_base)",
                "Sycl_TLA(00_padded)",
                "Sycl_TLA(00_sycl_q)",
                "Sycl_TLA(03_streamk)",
                "Sycl_TLA(03_dp)",
                "",
                "",
                "oneMKL(BLAS)",
                "oneDNN(matmul)",
                "PyTorch->oneDNN",
                "SYCL-TLA(XeTLA)",
                "oneMKL(BLAS)",
                "oneDNN(matmul)",
                "PyTorch->oneDNN",
                "SYCL-TLA(XeTLA)",
            ])
            sheet.append(["", 64, 4096, 4096, "Prefill", None, None, None, 100.0, None, None, None, 102.0, None, None, None, None, None, None, None, None, None, None])
            sheet.append(["", 128, 4096, 4096, "Prefill", None, None, None, 110.0, None, None, None, 112.0, None, None, None, None, None, None, 300.0, None, None, None])
            workbook.save(workbook_path)

            args = profiler.build_parser().parse_args(
                [
                    "--workspace",
                    tmpdir,
                    "--ali-workbook",
                    str(workbook_path),
                    "--max-shapes",
                    "1",
                    "--skip-run",
                    "--probe-mode",
                    "off",
                ]
            )
            outputs = profiler.workflow(args)

            shapes_doc = profiler.read_json(Path(outputs["workspace"]) / "inputs" / "gemm_target_shapes.json")
            reference_doc = profiler.read_json(Path(outputs["reference_doc"]))
            comparison = profiler.read_json(Path(outputs["reference_comparison"]))

        self.assertEqual(shapes_doc["shape_limit"], 1)
        self.assertEqual(shapes_doc["unlimited_shape_count"], 3)
        self.assertEqual(len(shapes_doc["shapes"]), 1)
        self.assertEqual(len(reference_doc["entries"]), 1)
        self.assertEqual(reference_doc["unlimited_reference_entries"], 3)
        self.assertEqual(comparison["summary"]["reference_entries"], 1)

    def test_negative_max_shapes_is_rejected(self):
        shapes_doc = profiler.default_shapes("bf16")

        with self.assertRaisesRegex(ValueError, "--max-shapes must be non-negative"):
            profiler.limit_shapes_and_reference(shapes_doc, max_shapes=-1)

    def test_ali_workbook_conflicts_with_explicit_shapes_or_reference(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ali.xlsx"
            shapes_path = Path(tmpdir) / "shapes.json"
            workbook = Workbook()
            workbook.save(workbook_path)
            shapes_path.write_text("{}", encoding="utf-8")
            args = profiler.build_parser().parse_args(
                [
                    "--workspace",
                    tmpdir,
                    "--ali-workbook",
                    str(workbook_path),
                    "--shapes-json",
                    str(shapes_path),
                ]
            )

            with self.assertRaisesRegex(ValueError, "mutually exclusive"):
                profiler.workflow(args)

    def test_build_reference_comparison_matches_dispatch_against_reference(self):
        dispatch_table = {
            "entries": [
                {
                    "shape_id": "rcr_bf16_8192_4096_4096",
                    "candidate_id": "cand_a",
                    "selected_metric": 150.0,
                }
            ]
        }
        reference_doc = {
            "dataset_id": "ali",
            "entries": [
                {
                    "shape_id": "rcr_bf16_8192_4096_4096",
                    "reference_provider": "Sycl_TLA(03_dp)",
                    "reference_tflops": 160.0,
                    "supported": True,
                }
            ],
        }

        comparison = profiler.build_reference_comparison(dispatch_table, reference_doc)

        self.assertEqual(comparison["summary"]["matched"], 1)
        self.assertEqual(comparison["entries"][0]["selected_candidate_id"], "cand_a")
        self.assertEqual(comparison["entries"][0]["selected_vs_reference_ratio"], 0.9375)

    def test_phase_a_summary_includes_anomaly_report(self):
        summary = profiler.build_phase_a_summary(
            {
                "probe_mode": "run",
                "hw_reference_spec_id": "bmg_g21",
                "dpas_baseline_probe": {"status": "pass", "avg_tflops": "1.23"},
                "compiler_flags_probe": {"results": [{"compiler_profile_id": "bmg.small_tile.default", "status": "pass"}]},
                "anomaly_report": {"anomalies": [{"candidate_id": "cand_large"}], "auto_block_rules": [{"rule_id": "rule.large"}]},
            },
            profiler.default_constraints(),
            [],
        )

        self.assertEqual(summary["hw_reference_spec_id"], "bmg_g21")
        self.assertEqual(summary["anomaly_report"]["anomalies"][0]["candidate_id"], "cand_large")


if __name__ == "__main__":
    unittest.main()
